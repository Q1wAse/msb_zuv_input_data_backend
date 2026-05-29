from typing import NamedTuple, Tuple, Any
from enum import Enum
from datetime import date
from pathlib import Path
import sys, os, io, openpyxl

from openpyxl.cell import Cell
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side

from urllib.parse import parse_qs
from decimal import Decimal
from flask import session, g, abort, send_file
import re

from sqlalchemy import inspect, text, exc
from sqlalchemy.orm.attributes import InstrumentedAttribute
from sqlalchemy.orm.exc import NoResultFound
from sqlalchemy.orm import scoped_session, sessionmaker

from msb_zuv_input_data_backend.config import Config
from msb_zuv_input_data_backend.database import engine_py, db_py

try:
    from access_control_center.centrilized_database_pool import get_session
except (ImportError, ModuleNotFoundError):
    get_session = None
#============================================================================================
#============================================================================================
class EnumMsg(Enum):
    SUCCESS         			    = 0
    SYSTEM_ERROR			        = 1
    INCORRECT_PARAM                 = 2
    INCORRECT_TAB_KEY               = 3
    INCORRECT_PATCH_INPUT_DATA      = 4
    INCORRECT_TEMPLATE_NAME         = 5
    NO_SELECTED_COLUMNS             = 6
    SETTINGS_FOR_REPORT_NOT_FOUND   = 7

#============================================================================================
#============================================================================================
msg_list = {
    EnumMsg.SUCCESS		                    : { 'code' : 200, 'is_err': False,	'msg' : 'Успешное выполнение операции' },
    EnumMsg.SYSTEM_ERROR			        : { 'code' : 500, 'is_err': True,	'msg' : 'Системная ошибка' },
    EnumMsg.INCORRECT_PARAM			        : { 'code' : 400, 'is_err': True,	'msg' : 'Неверно задано значение для %' },
    EnumMsg.INCORRECT_TAB_KEY		        : { 'code' : 400, 'is_err': True,	'msg' : 'Некорректное имя ключа таблицы' },
    EnumMsg.INCORRECT_PATCH_INPUT_DATA		: { 'code' : 400, 'is_err': True,	'msg' : 'Некорректный формат обновляемых данных' },
    EnumMsg.INCORRECT_TEMPLATE_NAME		    : { 'code' : 400, 'is_err': True,	'msg' : 'Некорректное наименование шаблона' },
    EnumMsg.NO_SELECTED_COLUMNS		        : { 'code' : 400, 'is_err': True,	'msg' : 'Должен быть выбран хотя бы один столбец' },
    EnumMsg.SETTINGS_FOR_REPORT_NOT_FOUND	: { 'code' : 400, 'is_err': True,	'msg' : 'Не удалось найти настройки для выбранных отчётов из таблицы: "tab_factories_d816_4"' },
}

TABLES_MAP = {
    'map_bs_product': {
        'tab_name': 'tab_map_bs_product_d816_4',
        'fields': 'id,name,id_product,koef,factory,type_raspr,sobstv,mest',
        'mutable' : True
    },
    'products': {
        'tab_name': 'tab_product_d816_4',
        'fields': 'id, name'
    },
    'factory': {
        'tab_name': 'tab_factory_d816_4',
        'fields': 'id, name'
    },
    'type_raspr': {
        'tab_name': 'tab_type_raspr_d816_4',
        'fields': 'id, name'
    },
    'sobstv': {
        'tab_name': 'tab_sobstv_d816_4',
        'fields': 'id, name'
    },
    'mest': {
        'tab_name': 'tab_mest_d816_4',
        'fields': 'id, name'
    },
    'category_product': {
        'tab_name': 'tab_view_category_product_d816_4',
        'fields': 'id, name'
    },
    'ost': {
        'tab_name': 'tab_ost_d816_4',
        'fields': 'id,tab_factory_d816_4_ids,tab_category_product_d816_4_ids,tab_product_d816_4_ids,value,value_korr',
        'mutable' : True
    },

    'factories' :{
        'tab_name': 'tab_factories_d816_4',
        'fields' : 'id,name'
    },
    'type_reports' :{
        'tab_name': 'tab_type_reports_d816_4',
        'fields' : 'id,name'
    },
    'data_type' :{
        'tab_name': 'tab_view_io_bcblm0003_d816_4',
        'fields' : 'id,name'
    },
    'vers_plan' :{
        'tab_name': 'tab_view_vers_plan_d816_4',
        'fields' : 'id,name'
    },
    'var_plan' :{
        'tab_name': 'tab_view_var_plan_d816_4',
        'fields' : 'tab_vers_plan_ids,name,id'
    }
}

template_list = [
        'Астрахань',
        'Сосногорск'
    ]
template_setups = [
        {  # Астраханский ГПЗ
            'index'         : 1,
            'template_name' : 'Астрахань',
            'do'            : 38,
            'pj'            : 7
        },
        {  # Сосногорский ГПЗ
            'index'         : 2,
            'template_name' : 'Сосногорск',
            'do'            : 38,
            'pj'            : 1
        }
    ]

main_folder = "/opt/foresight/msb_zuv_input_data_backend" if sys.platform.lower() in 'linux' else os.getcwd()
file_folder = "file"
sql_folder = "sql"

#============================================================================================
#============================================================================================
gv_collect_log_status = 0
gv_collect_log = ""

def clear_loc_log():
    global gv_collect_log_status
    global gv_collect_log

    gv_collect_log_status = 0
    gv_collect_log = ""#"\nn\nn"

def loc_log(msg):
    global gv_collect_log_status
    global gv_collect_log

    gv_collect_log_status = -1
    #gv_collect_log = gv_collect_log + "\nn" + msg
    gv_collect_log = gv_collect_log + " " + msg

def loc_log_new(func, locs, err):
    global gv_collect_log_status
    global gv_collect_log

    gv_collect_log_status = -1
    #gv_collect_log = gv_collect_log + "\nn" + msg

    # gv_collect_log = gv_collect_log + " " + msg
    gv_collect_log = gv_collect_log + " " +  "func::" + func + "::locs::" + str(locs) + "::" + str(err)

def get_db_connection():
    if Config.SERVERBASE_MODE == 'PYTHON':
        return db_py
    elif Config.SERVERBASE_MODE == 'WSGI':
        try:
            return get_session()
        except Exception as e:
            loc_log(str(e))
            abort(msg_list[EnumMsg.SYSTEM_ERROR].get('code'), description=get_msg_struct(EnumMsg.SYSTEM_ERROR)[0]['message'])

def is_valid_date(date_string):
    try:
        date.fromisoformat(date_string)
        return True
    except ValueError:
        return False

def get_validate_param(param, field_name):
    value = param.get(field_name)
    if field_name == "tab_id":
        is_valid = isinstance(value, str) and len(value) > 0
    elif field_name == "filter" and value is not None:
        is_valid = isinstance(value, str)
    elif field_name == "page":
        is_valid = isinstance(value, int) and value >= 1
    elif field_name == "limit":
        is_valid = isinstance(value, int) and (1 <= value <= 100)
    elif field_name == "year":
        is_valid = isinstance(value, int) and (0 <= value <= 9999)
    else:
        is_valid = True #(value is not None)

    if not is_valid:
        err_msg = get_msg_struct(EnumMsg.INCORRECT_PARAM, field_name)[0]['message']
        loc_log(err_msg)
        abort(msg_list[EnumMsg.INCORRECT_PARAM].get('code'), description=err_msg)
    return value

def is_msg_id_valid(msg_id):
    try:
        EnumMsg(msg_id)
        return True
    except ValueError:
        return False

def get_msg_struct(msg_id, value = ""):
    msg = 'неизвестная ошибка'
    enum_msg_local = EnumMsg(msg_id) if type(msg_id) is int else msg_id
    if is_msg_id_valid(msg_id):
        msg = msg_list[enum_msg_local].get('msg')
    type_msg = 'ошибки' if msg_list[enum_msg_local].get('is_err') else 'сообщения'
    msg = 'Код ' + type_msg + ': ' + str(enum_msg_local.value) + '. ' + msg + gv_collect_log
    msg = msg.replace('%',value)
    return {'message' : msg }, msg_list[enum_msg_local].get('code')

def exec_sql_from_file(file_name, params = {}):
    db = get_db_connection()
    if len(file_name) > 0:
        full = Path(main_folder) / Path(sql_folder) /  file_name
        if full.is_file():
            try:
                with open(full, 'r', encoding='utf-8') as file:
                    sql_exec = file.read()
                    query = db.execute(text(sql_exec), params)
                    db.commit()
                    return query
            except Exception as e:
                loc_log(str(e))
                abort(msg_list[EnumMsg.SYSTEM_ERROR].get('code'), description=get_msg_struct(EnumMsg.SYSTEM_ERROR)[0]['message'])
        return '-2'
    return '-1'

def get_param_connect():
    db = get_db_connection()
    return  db.execute(text("SELECT connection_params FROM tab_params_email_d314 WHERE system_code = 'IS_KAO_DATA'")).first()
#============================================================================================
def convert_row(row):
    result = {}
    for key, value in row.items():
        if isinstance(value, Decimal):
            result[key] = float(value) #str(value)
        else:
            result[key] = value
    return result
#============================================================================================
def get_pagin_data_old(v_tab_id, v_filter, v_page, v_limit):
    db = get_db_connection()
    offset = max(0, (v_page - 1) * v_limit)

    cond = "WHERE name ILIKE :filter" if v_filter else ""
    params = {
        "filter": f"%{v_filter}%" if v_filter else "%",
        "limit": v_limit,
        "offset": offset
    }

    # print(str(v_tab_id) + " " + str(dict(TABLES_MAP).values()))

    v_tabname = TABLES_MAP[v_tab_id].get('tab_name')
    field_list = TABLES_MAP[v_tab_id].get('fields')

    try:
        count_sql = text(f"SELECT count(*) FROM {v_tabname} {cond}")
        total = db.execute(count_sql, params).scalar()

        sql_text = text(f"""
            SELECT {field_list} FROM {v_tabname} 
            {cond} 
            LIMIT :limit OFFSET :offset
        """)
        rows = db.execute(sql_text, params).mappings() #fetchall()
        if rows:
            rows = [convert_row(row) for row in rows]
        # return [{'count' : total }, {'rows' : rows}]
        return {'count' : total, 'rows' : rows}

    except Exception as e:
        loc_log_new(sys._getframe(0).f_code.co_name, locals(), e)
        return [0, []]
#============================================================================================
#############################################################################################
#############################################################################################
#############################################################################################
#============================================================================================
def get_pagin_data(v_tab_id, v_filter, v_page, v_limit):
    db = get_db_connection()
    offset = max(0, (v_page - 1) * v_limit)

    cond = "WHERE name ILIKE :filter" if v_filter else ""
    params = {
        "filter": f"%{v_filter}%" if v_filter else "%",
        "limit": v_limit,
        "offset": offset
    }

    # print(str(v_tab_id) + " " + str(dict(TABLES_MAP).values()))

    v_tabname = TABLES_MAP[v_tab_id].get('tab_name')
    field_list = TABLES_MAP[v_tab_id].get('fields')

    try:
        count_sql = text(f"SELECT count(*) FROM {v_tabname} {cond}")
        total = db.execute(count_sql, params).scalar()

        sql_text = text(f"""
            SELECT {field_list} FROM {v_tabname} 
            {cond} 
            LIMIT :limit OFFSET :offset
        """)
        rows = db.execute(sql_text, params).mappings() #fetchall()
        if rows:
            rows = [convert_row(row) for row in rows]
        # return [{'count' : total }, {'rows' : rows}]
        return rows

    except Exception as e:
        loc_log_new(sys._getframe(0).f_code.co_name, locals(), e)
        return []
#============================================================================================
def patch_data(resource_key, data_list):
    db = get_db_connection()
    config = TABLES_MAP.get(resource_key)

    if not config or not isinstance(data_list, list):
        abort(msg_list[EnumMsg.INCORRECT_PATCH_INPUT_DATA].get('code'), description=get_msg_struct(EnumMsg.INCORRECT_PATCH_INPUT_DATA)[0]['message'])

    results = {
        "success"   : [],
        "errors"    : []
    }

    try:
        with db.begin():
            for item in data_list:
                id_record = item.get('id')
                if not id_record:
                    results["errors"].append(
                        {
                            "id"        : "unknown",
                            "message"   : "ID not found"
                        }
                    )
                    continue

                update_dict = {k: v for k, v in item.items() if v is not None and k != 'id'}

                if update_dict:
                    set_fields = ", ".join([f"{col} = :{col}" for col in update_dict.keys()])

                    sql = text(
                        f"""
                            UPDATE {config.get('tab_name')} 
                            SET {set_fields} 
                            WHERE id = :id
                        """
                    )

                    params = {
                        **update_dict,
                        "id": id_record
                    }
                    res = db.execute(sql, params)

                    if res.rowcount > 0:
                        results["success"].append(id_record)
                    else:
                        results["errors"].append(
                            {
                                "id"        : id_record,
                                "message"   : "Not found"
                            }
                        )

        return {
                "status"    : "completed",
                "details"   : results
        }
    except Exception as e:
        loc_log_new(sys._getframe(0).f_code.co_name, locals(), e)
        abort(msg_list[EnumMsg.SYSTEM_ERROR].get('code'), description=get_msg_struct(EnumMsg.SYSTEM_ERROR)[0]['message'])
#============================================================================================
def map_pg_to_frontend(pg_type):
    mapping = {
        'integer'                       : 'number',
        'numeric'                       : 'number',
        'real'                          : 'number',
        'double precision'              : 'number',
        'character varying'             : 'string',
        'text'                          : 'string',
        'boolean'                       : 'boolean',
        'timestamp without time zone'   : 'datetime',
        'date'                          : 'date'
    }
    return mapping.get(pg_type, 'string')
#============================================================================================
def get_struct_table(key_tab):
    db = get_db_connection()
    config = TABLES_MAP.get(key_tab)
    if not config:
        abort(msg_list[EnumMsg.INCORRECT_TAB_KEY].get('code'), description=get_msg_struct(EnumMsg.INCORRECT_TAB_KEY)[0]['message'])

    field_list = { col.strip() for col in TABLES_MAP.get(key_tab).get('fields').split(',') }

    col_sql = text("""
        SELECT
            column_name,
            data_type,
            numeric_precision, 
            numeric_scale,
            character_maximum_length as max_len 
        FROM information_schema.columns WHERE
            table_name = :t_name AND
            column_name IN :fields
    """)

    try:
        res = db.execute(col_sql,
                   {
                       't_name' : config.get('tab_name'),
                       'fields' : tuple(field_list)
                   }
        ).fetchall()

        res_dict = []
        for row in res:
            type_data = map_pg_to_frontend(row.data_type)
            if type_data == 'number':
                res_dict.append(
                    {
                        row.column_name : {
                            'type'      : type_data,
                            'precision' : row.numeric_precision,
                            'scale'     : row.numeric_scale
                        }
                    }
                )
            else:
                res_dict.append(
                    {
                        row.column_name: {
                            'type': type_data,
                            'len': row.max_len
                        }
                    }
                )

        return res_dict

    except Exception as e:
        loc_log_new(sys._getframe(0).f_code.co_name, locals(), e)
        abort(msg_list[EnumMsg.SYSTEM_ERROR].get('code'), description=get_msg_struct(EnumMsg.SYSTEM_ERROR)[0]['message'])
#============================================================================================
def get_row_list_msb_zuv_d816_4(year : int, ver_plan : int, var_plan : int, bs : list, do : int, pj : int, data_type : int):
    db = get_db_connection()
    col_sql = text("""
                SELECT
                    SUM(SUM),
                    BS,
                    CALYEAR, 
                    CALQUART,
                    CALMONTH 
                FROM tab_integ_get_preu_mirror_d816_4 WHERE
                    CALYEAR::INT = :year AND            -- Год планирования
                    BCBLM0001::INT = :ver_plan AND      -- Версия планирования
                    BCBLM0002::INT = :var_plan AND      -- Вариант планирования              
                    BS = ANY((:bs)::int[]) AND          -- Бюджетные статьи
                    BCBIM0002::INT = :do AND            -- Завод (Дочернее общество)
                    pj = :pj AND                        -- Перерабатывающий комплекс (Поставщики ЖУВ)
                    DATA_TYPE::INT = :data_type AND     -- Тип данных
                    CALMONTH <> 0 AND
                    DBS = 0
                GROUP BY BS, CALYEAR, CALQUART, CALMONTH
                ORDER by CALMONTH
            """)
    result = db.execute(col_sql,
                     {
                         'year'         : year,
                         'ver_plan'     : ver_plan,
                         'var_plan'     : var_plan,
                         'bs'           : f"{{{','.join(map(str, bs))}}}",
                         'do'           : do,
                         'pj'           : pj,
                         'data_type'    : data_type
                     }
                     ).fetchall()
    return result
#============================================================================================
def download_report(year, template_name):
    settings = [item for item in template_setups if item['template_name'] == template_name]
    if not settings:
        return

    if settings[0].get('all', False):
        settings = [item for item in template_setups if item['template_name'] != template_name]
    if not settings:
        return

    template_name += '.xlsx'
    path_template = str(Path(main_folder) / Path(file_folder) / Path(template_name))

    # Получаем рабочую книгу из шаблона
    wb = openpyxl.load_workbook(path_template)

    # Создаём буфер для наполнения
    buffer = io.BytesIO()

    named_value = wb.defined_names['_YEAR']
    if named_value:
        named_value.value = year

    #for index, item in enumerate(settings, start=1):
    for item in settings:
        do = item.get('do')
        pj = item.get('pj')
        index = item.get('index')

        def_range_bs = wb.defined_names[f'_BS{index}']
        if not def_range_bs:
            continue

        bs_col_index = -1
        dict_month_col_index = []
        sheet_name = ''
        for sheet_name_rng_bs, cell_coordinates in def_range_bs.destinations:
            bs_col_index, _, _, _ = range_boundaries(cell_coordinates)
            sheet_name = sheet_name_rng_bs
        for i in range(1, 13):
            def_range_month = wb.defined_names[f'_MONTH{index}_{i}']
            for sheet_name_rng_month, cell_coordinates in def_range_month.destinations:
                min_col, _, _, _ = range_boundaries(cell_coordinates)
                dict_month_col_index.append(min_col)

        if bs_col_index != -1 and sheet_name and len(dict_month_col_index) == 12:
            dict_bs = []
            sheet = wb[sheet_name]
            if not sheet:
                continue

            last_row = sheet.max_row
            for i in range(1, last_row + 1):
                cell = sheet.cell(row=i,column=bs_col_index)
                if cell.value is not None and str(cell.value).isdigit():
                    dict_bs.append(cell.value)
            if not dict_bs:
                continue

            query_plan_99_res = get_row_list_msb_zuv_d816_4(year, 22600, 2260099, dict_bs, do, pj, 1)
            query_plan_10_res = get_row_list_msb_zuv_d816_4(year, 22600, 2260010, dict_bs, do, pj, 1)
            query_fact_res = get_row_list_msb_zuv_d816_4(year, 0, 0, dict_bs, do, pj, 15)

            for i in range(1, last_row + 1):
                cell_bs = sheet.cell(row=i, column=bs_col_index)
                if cell_bs and cell_bs.value is not None and str(cell_bs.value).isdigit():
                    # for month_col in dict_month_col_index:
                    for ind, it in enumerate(dict_month_col_index, start=1):
                        for row in query_plan_99_res:
                            cell_val = sheet.cell(row=i, column=it)
                            if cell_val and cell_val.data_type == 'f':
                                continue
                            if int(row.bs) == int(cell_bs.value) and ind == row.calmonth:
                                cell_val.value = row.sum
                        for row in query_plan_10_res:
                            cell_val = sheet.cell(row=i, column=it+1)
                            if cell_val and cell_val.data_type == 'f':
                                continue
                            if int(row.bs) == int(cell_bs.value) and ind == row.calmonth:
                                cell_val.value = row.sum
                        for row in query_fact_res:
                            cell_val = sheet.cell(row=i , column=it+2)
                            if cell_val and cell_val.data_type == 'f':
                                continue
                            if int(row.bs) == int(cell_bs.value) and ind == row.calmonth:
                                cell_val.value = row.sum


    # Сохраняем подготовленные данные из шаблона
    wb.save(buffer)
    # Откатываем курсор в самое начало
    buffer.seek(0)
    # имя файла
    filename = template_name # "test_file.xlsx"

    return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
#============================================================================================
def get_data_from_query(sql_text, param=None):
    db = get_db_connection()
    if db:
        if param:
            return db.execute(text(sql_text), param).fetchall()
        else:
            return db.execute(text(sql_text)).fetchall()
    return []
#============================================================================================
def set_value_cell(cell, value, LevelTitle=0):
    simple_font = Font(
        name="Times New Roman",
        size=14,
        bold=True,
        color="000000"
    )
    formula_font = Font(
        name="Times New Roman",
        size=14,
        bold=True,
        color="0000FF"
    )
    title_font = {
        1 : Font(
                name="Times New Roman",
                size=10,
                bold=True
            ),
        2 : Font(
                name="Times New Roman",
                size=14,
                bold=True
            )
    }

    cell.value = value
    cell.number_format = '#,##0.00'

    if LevelTitle > 0:
        cell.font = title_font[LevelTitle]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        if cell.data_type == 'f':
            cell.font = formula_font
        else:
            cell.font = simple_font

def download_report2(selected_factories,selected_reports,columns):

    month = {
        1 : 'январь',
        2 : 'февраль',
        3 : 'март',
        4 : 'апрель',
        5 : 'май',
        6 : 'июнь',
        7 : 'июль',
        8 : 'август',
        9 : 'сентябрь',
        10 : 'октябрь',
        11 : 'ноябрь',
        12 : 'декабрь',
    }
    offset_ind_col = 1
    factories_all = [row.id for row in get_data_from_query("SELECT id FROM tab_factories_d816_4")]
    if not selected_factories:
        selected_factories = factories_all
    if not selected_reports:
        selected_reports = [1]

    settings = get_data_from_query(
        'SELECT id, "DO", pj FROM tab_factories_d816_4 WHERE id IN :factory_ids',
        {"factory_ids": tuple(selected_factories)})
    if not settings:
        return get_msg_struct(EnumMsg.SETTINGS_FOR_REPORT_NOT_FOUND)

    composite_keys_do_pj = [(row.DO, row.pj) for row in settings]

    bs_calc_mapping = get_data_from_query(
        """
                    SELECT "DO", pj, bs, calc FROM tab_bs_calc_map_d816_4 WHERE ("DO",pj) IN :composite_keys
                """,
        {"composite_keys": tuple(composite_keys_do_pj)})


    temp_data = {}
    for col in columns:
        type_id = int(col['typeData'])

        if type_id == 1:
            if type_id not in temp_data:
                temp_data[type_id] = []

            temp_data[type_id].append({ 'variant_planing': col.get('variantPlaning') })

        else:
            temp_data[type_id] = {'simple': True}

    columns_collect = [{key: val} for key, val in temp_data.items()]

    columns_text = []
    # for collect in columns_collect:
    for item in columns_collect:
        for type_id, collect in item.items():
            if type_id == 1:
                variants = [item['variant_planing'] for item in collect]

                query_res = get_data_from_query(
                    'SELECT id, name FROM tab_view_var_plan_d816_4 WHERE id IN :variant_ids',
                    {'variant_ids': tuple(variants)})
                for row in query_res:
                    columns_text.append({'typeData' : type_id, 'id' : row.id, 'name' : row.name})

            elif type_id == 15:
                query_res = get_data_from_query(
                    'SELECT id, name FROM tab_view_io_bcblm0003_d816_4 WHERE id = :type_id',
                    {'type_id': type_id})
                for row in query_res:
                    columns_text.append({'typeData' : type_id, 'id':row.id, 'name' : row.name})

    def get_txt_col(column):
        if 'typeData' in column:
            typeData = int(column.get('typeData'))
            if typeData == 1 and 'variantPlaning' in column:
                matched_item = next( (item for item in columns_text if int(item['id']) == int(column.get('variantPlaning'))), '' )
                return matched_item.get('name','')
            elif typeData == 15:
                matched_item = next((item for item in columns_text if int(item['id']) == typeData), '')
                return matched_item.get('name', '')
        return ''

    template_name = 'Астрахань.xlsx'
    path_template = str(Path(main_folder) / Path(file_folder) / Path(template_name))

    # Получаем рабочую книгу из шаблона
    wb = openpyxl.load_workbook(path_template)

    # Создаём буфер для наполнения
    buffer = io.BytesIO()

    #===================================================================================================================
    for factory_id in factories_all:
        loc_settings = [row for row in settings if row.id == factory_id]
        try:
            def_range_bs = wb.defined_names[f'_BS{factory_id}']
        except Exception as e:
            continue
        sheet_name_bs = ''
        bs_col_index = -1
        for sheet_name_rng_bs, cell_coordinates_rng_bs in def_range_bs.destinations:
            bs_col_index, _, _, _ = range_boundaries(cell_coordinates_rng_bs)
            sheet_name_bs = sheet_name_rng_bs
            sheet = wb[sheet_name_bs]
            if factory_id not in selected_factories:
                sheet.sheet_state = 'veryHidden'
                continue

            sheet.sheet_state = 'visible'
            if sheet and bs_col_index != -1:
                dict_bs = []
                last_row = sheet.max_row
                first_row = -1
                for i in range(1, last_row + 1):
                    cell = sheet.cell(row=i, column=bs_col_index)
                    if cell.value is not None and str(cell.value).isdigit():
                        dict_bs.append(cell.value)
                    if cell.value == 'ID':
                        first_row = i
                if not dict_bs or first_row == -1:
                    continue

                loc_bs_calc_mapping = [row for row in bs_calc_mapping if row.DO == loc_settings[0].DO and row.pj == loc_settings[0].pj]

                columns_layout = []
                for index_column, column in enumerate(columns):
                    columns_layout.append({
                        "type": f"year{column.get('dateRange')[0][-4:]}",
                        "data_type_col": index_column,
                        'IsNeedMerge' : True if index_column == 0 else False,
                        'col_name' : get_txt_col(column)
                    })
                for q in range(1, 5):
                    for index_column, column in enumerate(columns):
                        columns_layout.append({
                            "type": f"Q{q}",
                            "data_type_col":index_column,
                            'IsNeedMerge' : True if index_column == 0 else False,
                            'col_name' : get_txt_col(column)
                        })

                    for m in range(1, 4):
                        for index_column, column in enumerate(columns):
                            columns_layout.append({
                                "type": f"M{q}_{m}",
                                "data_type_col": index_column,
                                'IsNeedMerge' : True if index_column == 0 else False,
                                'col_name' : get_txt_col(column),
                                'calmonth': (q-1)*3+m
                            })

                count_columns = len(columns)

                # 'thin', 'medium', 'thick', 'double', 'dashed'
                thin_line = Side(border_style="thin", color="000000")
                full_border = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
                for row in sheet.iter_rows(
                        min_row=first_row,
                        max_row=last_row,
                        min_col=bs_col_index + 1 + offset_ind_col,
                        max_col=bs_col_index + 1 + offset_ind_col + len(columns_layout)-1):
                    for cell in row:
                        cell.border = full_border

                sheet.row_dimensions[first_row+1].height = 40

                for idx, col in enumerate(columns_layout):
                    col_num = idx + bs_col_index + 1 + offset_ind_col
                    col_letter = get_column_letter(col_num)
                    sheet.column_dimensions[col_letter].width = 22 #16.29
                    cell = sheet.cell(row=first_row+1, column=col_num)
                    set_value_cell(cell,col.get('col_name'), 1)

                    if col["IsNeedMerge"]:
                        sheet.merge_cells(start_row=first_row, start_column=col_num, end_row=first_row,
                                          end_column=col_num + count_columns - 1)
                        cell = sheet.cell(row=first_row, column=col_num)

                        if "year" in col["type"]:
                            set_value_cell(cell,col.get('type')[-4:], 2)
                        elif "Q" in col['type']:
                            q_number = col['type'][1]
                            set_value_cell(cell,f"{q_number} квартал", 2)
                        elif "M" in col["type"]:
                            set_value_cell(cell,month[col.get('calmonth')], 2)


                query_res_for_column = []
                # for column in columns:
                for index_column, column in enumerate(columns, start=1):
                    year = column.get('dateRange')[0][-4:]
                    ver_plan = column.get('versionPlaning', 0)
                    var_plan = column.get('variantPlaning', 0)
                    do = settings[0].DO
                    pj = settings[0].pj
                    data_type = column.get('typeData')
                    query_res_for_column.append(get_row_list_msb_zuv_d816_4(year, ver_plan, var_plan, dict_bs, do, pj, data_type))

                fill_obj_column(offset_ind_col, sheet,dict_bs,bs_col_index,first_row,last_row,loc_settings,loc_bs_calc_mapping,count_columns, columns_layout, query_res_for_column)
    #===================================================================================================================

    # Сохраняем подготовленные данные из шаблона
    wb.save(buffer)
    # Откатываем курсор в самое начало
    buffer.seek(0)
    # имя файла
    filename = "test_file.xlsx"

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
#============================================================================================
def fill_obj_column(offset_ind_col, sheet,dict_bs,bs_col_index,first_row,last_row,settings, bs_calc_mapping,count_sel_column, columns_layout, query_res_for_column):

    def get_col_letter_by_type(col_type, data_type_col):
        for idx, col in enumerate(columns_layout):
            if col["type"] == col_type and col["data_type_col"] == data_type_col:
                return get_column_letter(bs_col_index + idx + 1 + offset_ind_col)
        return None

    def set_cell_val(cell,col_letter,query_res,bs_calc_mapping):
        for map in bs_calc_mapping:
            if map.bs == cell_bs.value and map.calc != None and map.calc != '':
                parts = re.split(r'(\d+)', map.calc)
                src_formula = "="
                for p in parts:
                    if p.isdigit():
                        for i in range(1, last_row + 1):
                            loc_cell_bs = sheet.cell(row=i, column=bs_col_index)
                            if loc_cell_bs.value == p:
                                p = f"{col_letter}{i}"
                                break
                    src_formula += p
                set_value_cell(cell, src_formula)
                return
        for row in query_res:
            if int(row.bs) == int(cell_bs.value) and col["calmonth"] == row.calmonth:
                set_value_cell(cell, row.sum)
                return
        return

    for i in range(1, last_row + 1):
        cell_bs = sheet.cell(row=i, column=bs_col_index)
        if cell_bs and cell_bs.value is not None and str(cell_bs.value).isdigit():
            for idx, col in enumerate(columns_layout):
                col_num = idx + bs_col_index + 1 + offset_ind_col
                col_letter = get_column_letter(col_num)
                data_type_col = col["data_type_col"]

                if "year" in col["type"]:
                    q1 = get_col_letter_by_type("Q1", data_type_col)
                    q2 = get_col_letter_by_type("Q2", data_type_col)
                    q3 = get_col_letter_by_type("Q3", data_type_col)
                    q4 = get_col_letter_by_type("Q4", data_type_col)
                    # sheet[f"{col_letter}{i}"] = f"={q1}{i}+{q2}{i}+{q3}{i}+{q4}{i}"
                    set_value_cell(sheet[f"{col_letter}{i}"], f"={q1}{i}+{q2}{i}+{q3}{i}+{q4}{i}")

                elif "Q" in col["type"]:
                    q_num = col["type"][1]

                    m1 = get_col_letter_by_type(f"M{q_num}_1", data_type_col)
                    m2 = get_col_letter_by_type(f"M{q_num}_2", data_type_col)
                    m3 = get_col_letter_by_type(f"M{q_num}_3", data_type_col)
                    # sheet[f"{col_letter}{i}"] = f"={m1}{i}+{m2}{i}+{m3}{i}"
                    set_value_cell(sheet[f"{col_letter}{i}"], f"={m1}{i}+{m2}{i}+{m3}{i}")

                elif "M" in col["type"]:
                    cell_val = sheet[f"{col_letter}{i}"]
                    set_cell_val(cell_val,col_letter,query_res_for_column[col['data_type_col']],bs_calc_mapping)

    #    ГОД 2026               квартал 1               январь                  февраль                   март                  квартал 2               апрель                  май
    # [1     2       3]     [4       5       6]     [7      8       9]     [10     11      12]     [13     14      15]     [16     17      18]     [19     20      21]     [22     23      24]
    #                                               [1      2       3]     [4       5       6]     [7       8       9]                             [10     11      12]     [13     14      15]

    return False
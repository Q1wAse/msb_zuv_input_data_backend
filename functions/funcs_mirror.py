from typing import List, Dict, Any, Union
import sys, os, re, io, copy, traceback, openpyxl
from datetime import datetime
from psycopg2 import errors

from flask import session, g, abort, send_file
from sqlalchemy.exc import SQLAlchemyError, DataError, OperationalError
from sqlalchemy import text
from werkzeug.datastructures import FileStorage
from collections import defaultdict
from enum import Enum
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils.cell import range_boundaries, coordinate_to_tuple
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.styles import Border, Side, NamedStyle
# from openpyxl.styles import Font, Alignment, numbers, Border, Side, PatternFill
# from openpyxl.formatting.rule import CellIsRule

import msb_zuv_input_data_backend.functions.utility_functions as uf
import msb_zuv_input_data_backend.functions.formula_translator as formula_translator


# ============================================================================================
class EnumCellType(Enum):
    TITLE_LVL0 = 0
    TITLE_LVL1 = 1
    TITLE_GROUP1_LVL1 = 2
    TITLE_GROUP1_LVL2 = 3
    INPUT = 4 #TODO Обычный шрифт. Название осталось старое
    SIMPLE_FORMULA = 5
    FONT_CHECKER1 = 6
    PERCENT = 7
    POSITIVE_NEGATIVE = 8
    SECOND_MINUS_FIRST = 9
    PERCENT_OF_COMPLETE = 10
    PERCENT_OF_OUTPUT = 11
    CELL_INPUT = 12


class EnumColumnSettings(str, Enum):
    KEY_BS = 'key_bs'
    KEY_SPEC = 'key_spec'
    KEY_INPUT = 'key_input'
    FORMULA_MONTH = 'formula_month'
    FORMULA_POO = 'formula_PercentOfOutput'
    ROW_CHECK = 'row_check'

class EnumInternalKeyPatterns(str, Enum):
    MONTH = r".*M(\d+)_(\d+)"

class EnumFuncModuParameter(int, Enum):
    """
    1 - [Реализация] Доходы от реализации продуктов и нефтегазопереработки
      form_1_034
    2 - [Производство, переработка] Объем переработки сырья и производства продукции
      form_1_003
    3 - [Добыча] Объем добычи углеводородов
      form_1_040
      form_1_340
    4 - [Остатки] Материальный баланс движения продукции ПАО "Газпром"
      form_1_406
    """
    realisation = 1
    proizv_pererab = 2
    dobicha = 3
    ost = 4
# ============================================================================================

# g_report_template_name = "Астрахань.xlxs"
g_report_template_name = "мсб свод (общий).xlsx"

# FONT_SIMPLE = Font(name="Times New Roman", size=14, bold=True, color="000000")
# FONT_FORMULA = Font(name="Times New Roman", size=14, bold=True, color="0000FF")
# FONT_TITLE_LVL0 = Font(name="Times New Roman", size=10, bold=True)
# FONT_TITLE_LVL1 = Font(name="Times New Roman", size=14, bold=True)
# FONT_GREEN = Font(name="Times New Roman", size=14, color='00B050')
# FONT_RED = Font(name="Times New Roman", size=14, color='C00000')
#
# ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
#
# RULE_POSITIVE = CellIsRule(operator='greaterThan', formula=['0'], font=FONT_GREEN)
# RULE_NEGATIVE = CellIsRule(operator='lessThan', formula=['0'], font=FONT_RED)
#
# THIN_LINE = Side(border_style="thin", color="000000")
# FULL_BORDER = Border(left=THIN_LINE, right=THIN_LINE, top=THIN_LINE, bottom=THIN_LINE)

G_STYLE_FONT_SIMPLE = None
G_STYLE_FONT_FORMULA = None
G_STYLE_FONT_TITLE_LVL0 = None
G_STYLE_FONT_TITLE_LVL1 = None
G_STYLE_RULE_GREEN_RED_DASH = None
G_STYLE_FULL_BORDER = None
G_STYLE_FULL_BORDER2 = None
G_STYLE_FORMAT_PERCENTAGE = None
G_STYLE_FONT_TITLE_LVL1_GROUP1 = None
G_STYLE_FONT_TITLE_LVL2_GROUP1 = None
G_STYLE_FONT_CHECKER1 = None
G_STYLE_RULE_DASH_FOR_ZERO = None
G_STYLE_CELL_INPUT = None

SHEETS_SETTINGS = []
SHEETS_SETTINGS_REPORT_LIST = []
SHEETS_SETTINGS_FACTORY_LIST = []

template_list = [
    'Астрахань',
    'Сосногорск'
]
template_setups = [
    {  # Астраханский ГПЗ
        'index': 1,
        'template_name': 'Астрахань',
        'do': 38,
        'pj': 7
    },
    {  # Сосногорский ГПЗ
        'index': 2,
        'template_name': 'Сосногорск',
        'do': 38,
        'pj': 1
    }
]

generating_report_settings = {
    # ключ в "Таблица Типы отчётов" | index - индекс именованного диапазона
    '19': {'index': '1'},  # План общий
    '20': {'index': '4'},  # Факт общий
    '21': {'index': '2'},  # Баланс ЗС
    '23': {'index': '3'},  # ЕЖО
    '24': {'index': '5'},  # 1-003-О
}


# ============================================================================================
def get_sheets_settings():
    return uf.get_dict_data_from_query("SELECT * FROM tab_view_sheet_id_list_d816_4")
def get_report_sheet_id_list(list, type_generation):
    if list:
        return [
            item.get('sheet_id')
            for item in list
            if item.get('type_generation') == type_generation
        ]
    return []
def get_sheet_list_by_field(list, field_name, field_filter):
    if list:
        return [
            item.get('sheet_id')
            for item in list
            if item.get(field_name) == field_filter
        ]
    return []
def get_sheet_list_by_fields(
        items_list: List[Dict[str, Any]],
        filters: Dict[str, Any],
        output_fields: Union[List[str], Dict[str, str]] = None
) -> List[Dict[str, Any]]:
    """
    Получить список лисов с заданным условием и перечисленным полями, которые можно переименовать
    """
    if not items_list:
        return []

    if output_fields is None:
        output_fields = {'sheet_id': 'sheet_id'}

    if isinstance(output_fields, list):
        fields_mapping = {field: field for field in output_fields}
    else:
        fields_mapping = output_fields

    matched_items = []
    for item in items_list:
        if all(item.get(field) == value for field, value in filters.items()):
            extracted_data = {}
            for original_name, custom_name in fields_mapping.items():
                extracted_data[custom_name] = item.get(original_name)
            matched_items.append(extracted_data)
    return matched_items
# ============================================================================================
def get_sheet_list_with_this_open(list_data, selected_reports):
    sheet_id_list = []
    if list_data:
        if selected_reports:
            open_with_this_sheet_id_list = [
                sheet_data.get('open_with_this_sheet_id')
                for sheet_data in list_data
                for item in selected_reports
                if sheet_data.get('sheet_id') == int(item) and sheet_data.get('type_sheet') == 'REPORT'
            ]
            if open_with_this_sheet_id_list:
                for sheet_data in open_with_this_sheet_id_list:
                    if sheet_data:
                        sheet_id = [int(number_str) for number_str in re.findall(r'\d+', sheet_data)]
                        if sheet_id:
                            sheet_id_list = list(set(sheet_id_list) | set(sheet_id))
    return sheet_id_list
# ============================================================================================
def get_row_list_msb_zuv_d816_4(
        year: int,
        ver_plan: int,
        var_plan: int,
        bs: list,
        key_input: list,
        do: int,
        pj: int,
        data_type: int,
        sheet_id: int,
        only_year: bool = False
):
    period = 'CALMONTH = 0 AND CALQUART = 0 AND' if only_year else 'CALMONTH <> 0 AND'
    db = uf.get_db_connection()
    col_sql = text(f"""
                    SELECT
                        SUM(SUM),
                        BS,
                        CALYEAR, 
                        CALQUART,
                        CALMONTH,
                        sheet_id
                    FROM tab_integ_get_preu_mirror_d816_4 WHERE
                        CALYEAR::INT = :year AND            -- Год планирования
                        BCBLM0001::INT = :ver_plan AND      -- Версия планирования
                        BCBLM0002::INT = :var_plan AND      -- Вариант планирования              
                        BS = ANY((:bs)::int[]) AND          -- Бюджетные статьи
                        BCBIM0002::INT = :do AND            -- Завод (Дочернее общество)
                        pj = :pj AND                        -- Перерабатывающий комплекс (Поставщики ЖУВ)
                        DATA_TYPE::INT = :data_type AND     -- Тип данных
                        {period}
                        DBS = 0 AND
                        sheet_id = 0
                        --sheet_id is null
                    GROUP BY BS, CALYEAR, CALQUART, CALMONTH, sheet_id
                
                UNION ALL
                
                    SELECT
                        SUM(SUM),
                        BS,
                        CALYEAR, 
                        CALQUART,
                        CALMONTH,
                        sheet_id
                    FROM tab_integ_get_preu_mirror_d816_4 WHERE
                        CALYEAR::INT = :year AND            -- Год планирования
                        BCBLM0001::INT = :ver_plan AND      -- Версия планирования
                        BCBLM0002::INT = :var_plan AND      -- Вариант планирования              
                        BS = ANY((:key_input)::int[]) AND   -- Сгенерированные ключи для сохранения введённых данных
                        -- BCBIM0002::INT = :do AND            -- Завод (Дочернее общество)
                        -- pj = :pj AND                        -- Перерабатывающий комплекс (Поставщики ЖУВ)
                        DATA_TYPE::INT = :data_type AND     -- Тип данных
                        {period}
                        DBS = 0 AND
                        sheet_id = :_sheet_id
                    GROUP BY BS, CALYEAR, CALQUART, CALMONTH, sheet_id
                
                ORDER by CALMONTH
            """)
    result = db.execute(col_sql,
                        {
                            'year': year,
                            'ver_plan': ver_plan,
                            'var_plan': var_plan,
                            'bs': f"{{{','.join(map(str, bs))}}}",
                            'key_input': f"{{{','.join(map(str, key_input))}}}",
                            'do': do,
                            'pj': pj,
                            'data_type': data_type,
                            '_sheet_id' : sheet_id,
                        }
                        ).fetchall()
    return result

# ============================================================================================
# ============================================================================================
def get_hard_mirror_pattern():
    return r'(?P<CELL_OFFSET>\{[^}]+\})' \
           r'|(?P<ESCAPED>\$[0-9]+)' \
           r'|(?P<FUNC_OR_VAR>[a-zA-Zа-яА-ЯёЁ0-9_]+)' \
           r'|(?P<OPERATOR>[;(),+\-*/:])' \
           r'|(?P<NUMBER>[0-9]+)'


def get_data_from_named_range_name(wb, name):
    Exec = False
    named_range = None
    sheet = None
    min_col, min_row, max_col, max_row = 0, 0, 0, 0
    sheet_name = ''
    try:
        named_range = wb.defined_names[name]
        for sheet_name, cell_coordinates in named_range.destinations:
            min_col, min_row, max_col, max_row = range_boundaries(cell_coordinates)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        Exec = True
    except Exception as e:
        try:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if name in sheet.defined_names:
                    named_range = sheet.defined_names[name]

                    for sheet_name2, cell_coordinates in named_range.destinations:
                        min_col, min_row, max_col, max_row = range_boundaries(cell_coordinates)

                    Exec = True
                    break
        except Exception as e:
            return {'Exec': False}
    return {
        'Exec': True,
        'sheet_name': sheet_name,
        'sheet': sheet,
        'min_col': min_col,
        'min_row': min_row,
        'max_col': max_col,
        'max_row': max_row
    } if Exec else {'Exec': False}
def get_data_from_named_range(wb, named_range):
    sheet = None
    min_col, min_row, max_col, max_row = 0, 0, 0, 0
    sheet_name = ''
    try:
        for sheet_name, cell_coordinates in named_range.destinations:
            min_col, min_row, max_col, max_row = range_boundaries(cell_coordinates)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        if sheet_name == '':
            sheet = wb[wb.sheetnames[named_range.localSheetId]]
            sheet_name = sheet.title
    except Exception as e:
        return {'Exec': False}
    return {
        'Exec': True,
        'sheet_name': sheet_name,
        'sheet': sheet,
        'min_col': min_col,
        'min_row': min_row,
        'max_col': max_col,
        'max_row': max_row
    }
def get_named_range_from_sheet_id_and_nr_name(wb, sheet_id, partial_named_range_name):
    sheet = None
    min_col, min_row, max_col, max_row = 0, 0, 0, 0
    sheet_name = ''
    try:
        sheet = get_sheet_from_sheet_id(wb, sheet_id)
        sheet_name = sheet.title
        for item in wb.defined_names:
            if partial_named_range_name in item:
                named_range = wb.defined_names[item]
                for sheet_name2, cell_coordinates in named_range.destinations:
                    min_col, min_row, max_col, max_row = range_boundaries(cell_coordinates)
                if sheet_name2 == sheet_name:
                    return named_range

    except Exception as e:
        return None
    return None
def get_named_rng_partial_name(wb, partial_name):
    """
        Возвращает именованный диапазон по первому вхождению его имени из списка именованных диапазонов
    """
    # поиск в зоне видимости КНИГА
    for item in wb.defined_names:
        if partial_name in item:
            return wb.defined_names[item]
    # поиск в зоне видимости ЛИСТ
    for sheet_name in wb.sheetnames:
        for item in wb[sheet_name].defined_names:
            if partial_name in item:
                return wb[sheet_name].defined_names[item]
    return None
def get_sheet_id_from_sheet(sheet):
    sheet_name = sheet.title
    try:
        for item in sheet.defined_names:
            if '_SHEET_ID' in item:
                return item.replace('_SHEET_ID','')
    except Exception as e:
        pass
    return None
def get_sheet_from_sheet_id(wb,sheet_id):
    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet:
                for item in sheet.defined_names:
                    if f'_SHEET_ID{sheet_id}' == item:
                        return sheet
    except Exception as e:
        pass
    return None
def get_common_column_settings(start_col_index):
    return {
                EnumColumnSettings.KEY_BS: start_col_index + 0,
                EnumColumnSettings.KEY_SPEC: start_col_index + 1,
                EnumColumnSettings.KEY_INPUT: start_col_index + 2,
                EnumColumnSettings.FORMULA_MONTH: start_col_index + 3,
                EnumColumnSettings.FORMULA_POO: start_col_index + 4,
                EnumColumnSettings.ROW_CHECK: start_col_index + 5
            }

def SplitInternalKey(value):
    try:
        parts = str(value).split(':')[::-1]
        var_planing = parts[0] if str(parts[0]).lower() != 'none' else '0'
        data_type = parts[1]
        quarter = -1
        month = -1
        if len(parts) == 3:
            year = parts[2].replace('year','')
        else:
            year = parts[3].replace('year','')
            if parts[2][0] == "Q":
                quarter = int(parts[2][1])
            elif parts[2][0] == "M":
                match = re.search(EnumInternalKeyPatterns.MONTH, parts[2])
                if match:
                    quarter = int(match.group(1))
                    month = int(match.group(2))
        return {
            'var_planing' : var_planing,
            'data_type' : data_type,
            'quarter' : quarter,
            'month' : month,
            'year' : year,
        }
    except Exception as e:
        return None
    return None
def build_dict_row(data, parameter:int):
    # parameter:
    # 1 - [Реализация] Доходы от реализации продуктов и нефтегазопереработки
    #   form_1_034
    # 2 - [Производство, переработка] Объем переработки сырья и производства продукции
    #   form_1_003
    # 3 - [Добыча] Объем добычи углеводородов
    #   form_1_040
    #   form_1_340
    # 4 - [Остатки] Материальный баланс движения продукции ПАО "Газпром"
    #   form_1_406

    # 'bs': cell_key_input.value,
    # 'bcblm0002': var_planing,
    # 'data_type': data_type,
    # 'calquart': quarter,
    # 'calmonth': month,
    # 'calyear': year,

    # row = ''
    # row = f'{row}{data.get("year")},' # calyear
    # row = f'{row}{data.get("var_planing")[:-2] or 0},' # bcblm0001 (Версия планирования)
    # row = f'{row}{data.get("var_planing") or 0},' # bcblm0002 (Вариант планирования)
    # row = f'{row}{data.get("quarter") or 0},' # calquart
    # row = f'{row}{data.get("month") or 0},' # calmonth
    # row = f'{row}{data.get("data_type") or 0},' # data_type
    # row = f'{row}{0},' # parameter
    # row = f'{row}{0},' # sum
    # row = f'{row}{data.get("bs") or 0},' # bs
    # row = f'{row}{0},' # dbs
    # row = f'{row}{0},' # pj (Перерабатывающий комплекс)
    # row = f'{row}{0},' # syr
    # row = f'{row}{0},' # bcbim0002 (Завод)
    # row = f'{row}{0},' # sheet_id (Ключ листа из шаблона-xls)

    # return f'({row})'

    dict_row = {
        'calyear' : data.get("year"), # calyear
        'bcblm0001' : data.get("var_planing")[:-2] or 0, # bcblm0001 (Версия планирования)
        'bcblm0002' : data.get("var_planing") or 0, # bcblm0002 (Вариант планирования)
        'calquart' : data.get("quarter") or 0, # calquart
        'calmonth' : data.get("month") or 0, # calmonth
        'data_type' : data.get("data_type") or 0, # data_type
        'parameter' : parameter, # parameter
        'sum' : data.get("sum"), # sum
        'bs' : data.get("bs") or 0, # bs
        'dbs' : 0, # dbs
        'pj' : 0, # pj (Перерабатывающий комплекс)
        'syr' : 0, # syr
        'bcbim0002' : "000000000000", # bcbim0002 (Завод)
        'sheet_id' : data.get("sheet_id"), # sheet_id (Ключ листа из шаблона-xls)
    }
    return dict_row
def set_value_cell(cell, value, ColumnType: EnumCellType = EnumCellType.INPUT):
    global G_STYLE_FONT_SIMPLE, \
        G_STYLE_FONT_FORMULA, \
        G_STYLE_FONT_TITLE_LVL0, \
        G_STYLE_FONT_TITLE_LVL1, \
        G_STYLE_FONT_GREEN, \
        G_STYLE_FONT_RED, \
        G_STYLE_RULE_GREEN_RED_DASH, \
        G_STYLE_FULL_BORDER, \
        G_STYLE_FULL_BORDER2, \
        G_STYLE_FORMAT_PERCENTAGE, \
        G_STYLE_FONT_TITLE_LVL1_GROUP1, \
        G_STYLE_FONT_TITLE_LVL2_GROUP1, \
        G_STYLE_FONT_CHECKER1, \
        G_STYLE_RULE_DASH_FOR_ZERO, \
        G_STYLE_CELL_INPUT

    cell.value = value
    # cell.style = "_USER_CUSTOM_BORDER"

    if ColumnType in (EnumCellType.TITLE_LVL0, EnumCellType.TITLE_LVL1):
        if ColumnType == EnumCellType.TITLE_LVL0:
            chosen_style = {**G_STYLE_FONT_TITLE_LVL0}
        else:
            chosen_style = {**G_STYLE_FONT_TITLE_LVL1}

    elif ColumnType == EnumCellType.INPUT:
        if cell.data_type == 'f':
            chosen_style = {**G_STYLE_FONT_FORMULA}
        else:
            chosen_style = {**G_STYLE_FONT_SIMPLE}

    elif ColumnType == EnumCellType.SIMPLE_FORMULA:
        chosen_style = {**G_STYLE_FONT_FORMULA}

    elif ColumnType in (EnumCellType.PERCENT, EnumCellType.POSITIVE_NEGATIVE):
        chosen_style = {**G_STYLE_FONT_FORMULA}

        # cell.parent.conditional_formatting.add(cell.coordinate, RULE_POSITIVE)
        # cell.parent.conditional_formatting.add(cell.coordinate, RULE_NEGATIVE)

        if ColumnType == EnumCellType.PERCENT:
            chosen_style['number_format'] = G_STYLE_FORMAT_PERCENTAGE  # тип данных строка
    elif ColumnType == EnumCellType.TITLE_GROUP1_LVL1:
        chosen_style = {**G_STYLE_FONT_TITLE_LVL1_GROUP1}
    elif ColumnType == EnumCellType.TITLE_GROUP1_LVL2:
        chosen_style = {**G_STYLE_FONT_TITLE_LVL2_GROUP1}
    elif ColumnType == EnumCellType.FONT_CHECKER1:
        chosen_style = {**G_STYLE_FONT_CHECKER1}
    elif ColumnType == EnumCellType.CELL_INPUT:
        chosen_style = {**G_STYLE_CELL_INPUT}
    else:
        chosen_style = {**G_STYLE_FONT_SIMPLE}

    cell.font = chosen_style['font']
    cell.alignment = chosen_style['alignment']
    cell.fill = chosen_style['fill']
    cell.number_format = chosen_style['number_format']


# #============================================================================================
def extract_cell_styles(src_cell):
    return {
        'font': copy.copy(src_cell.font),
        'alignment': copy.copy(src_cell.alignment),
        'fill': copy.copy(src_cell.fill),
        'number_format': src_cell.number_format
    }


# def extract_optimal_border(source_cell):
#     b = source_cell.border
#
#     def clean_side(side):
#         if side and side.style:
#             return Side(style=side.style, color=side.color)
#         return None
#
#     return Border(
#         left=clean_side(b.left),
#         right=clean_side(b.right),
#         top=clean_side(b.top),
#         bottom=clean_side(b.bottom),
#         diagonal=clean_side(b.diagonal),
#         diagonal_direction=b.diagonal_direction,
#         outline=b.outline,
#         vertical=clean_side(b.vertical),
#         horizontal=clean_side(b.horizontal)
#     )
def extract_fast_border(source_cell):
    b = source_cell.border

    def clean_side(side):
        if side and side.style:
            return Side(style=side.style, color=side.color)
        return None

    return Border(
        left=clean_side(b.left),
        right=clean_side(b.right),
        top=clean_side(b.top),
        bottom=clean_side(b.bottom),
        diagonal=clean_side(b.diagonal),
        diagonal_direction=b.diagonal_direction,
        outline=b.outline,
        vertical=clean_side(b.vertical),
        horizontal=clean_side(b.horizontal)
    )


# def download_report2(download_type, selected_factories, selected_reports, src_columns):
def main_download_report(download_type, selected_factories, selected_reports, src_columns):
    global G_STYLE_FONT_SIMPLE, \
        G_STYLE_FONT_FORMULA, \
        G_STYLE_FONT_TITLE_LVL0, \
        G_STYLE_FONT_TITLE_LVL1, \
        G_STYLE_FONT_GREEN, \
        G_STYLE_FONT_RED, \
        G_STYLE_RULE_GREEN_RED_DASH, \
        G_STYLE_FULL_BORDER, \
        G_STYLE_FULL_BORDER2, \
        G_STYLE_FORMAT_PERCENTAGE, \
        G_STYLE_FONT_TITLE_LVL1_GROUP1, \
        G_STYLE_FONT_TITLE_LVL2_GROUP1, \
        G_STYLE_FONT_CHECKER1, \
        G_STYLE_RULE_DASH_FOR_ZERO, \
        G_STYLE_CELL_INPUT

    selected_factories = [int(item) for item in selected_factories]
    selected_reports = [int(item) for item in selected_reports]

    storage_sheet = defaultdict(lambda: defaultdict(dict))

    SHEETS_SETTINGS = get_sheets_settings()
    if not SHEETS_SETTINGS:
        return uf.get_msg_struct(uf.EnumMsg.SETTINGS_FOR_REPORT_NOT_FOUND)

    SHEETS_SETTINGS_REPORT_LIST = [item for item in SHEETS_SETTINGS if item.get('type_sheet') == 'REPORT']
    SHEETS_SETTINGS_FACTORY_LIST = [item for item in SHEETS_SETTINGS if item.get('type_sheet') == 'FACTORY']

    month = {
        1: 'январь',
        2: 'февраль',
        3: 'март',
        4: 'апрель',
        5: 'май',
        6: 'июнь',
        7: 'июль',
        8: 'август',
        9: 'сентябрь',
        10: 'октябрь',
        11: 'ноябрь',
        12: 'декабрь',
    }
    # ==================================================================================================================
    len_src_columns = len(src_columns)
    # ==================================================================================================================
    offset_ind_col = 1
    # factories_all = [str(row.id) for row in uf.get_data_from_query("SELECT id FROM tab_factories_d816_4")]
    factories_all = get_sheet_list_by_field(SHEETS_SETTINGS,'type_sheet','FACTORY')
    if not selected_factories:
        # selected_factories = factories_all
        selected_factories = get_sheet_list_with_this_open(SHEETS_SETTINGS, selected_reports)
    # reports_all = [str(row.id) for row in uf.get_data_from_query("SELECT id FROM tab_type_reports_d816_4")]
    reports_all = get_sheet_list_by_field(SHEETS_SETTINGS,'type_sheet','REPORT')
    # if not selected_reports:
    #     selected_reports = [1]

    settings = uf.get_data_from_query(
        'SELECT id, "DO", pj FROM tab_factories_d816_4 WHERE id IN :factory_ids',
        {"factory_ids": tuple(factories_all)})
    if not settings:
        return uf.get_msg_struct(uf.EnumMsg.SETTINGS_FOR_REPORT_NOT_FOUND)

    generating_type = {
        'NeedGeneratingReport': bool(selected_factories or
                                     (any(item in selected_reports for item in get_report_sheet_id_list(SHEETS_SETTINGS_REPORT_LIST,'GENERATING')))),
        'NeedStaticReport': any(item in selected_reports for item in get_report_sheet_id_list(SHEETS_SETTINGS_REPORT_LIST,'STATIC')),
    }

    if len_src_columns > 1:
        have_plan = 0
        have_fact = 0
        for col in src_columns:
            type_data = col.get('typeData', None)
            # План
            if type_data == '1':
                have_plan = 1
            # Факт
            if type_data == '15':
                have_fact = 1
        # Проверяем, что в списке выбранного среза есть факт и план
        if have_fact + have_plan != 2:
            generating_type['NeedStaticReport'] = False
    else:
        generating_type['NeedStaticReport'] = False

    def get_txt_col(column):
        if 'ColumnType' in column:
            if column.get('ColumnType') == 'Selected':
                if 'typeData' in column:
                    typeData = int(column.get('typeData'))
                    if typeData == 1 and 'variantPlaning' in column:
                        matched_item = next(
                            (item for item in columns_text if int(item['id']) == int(column.get('variantPlaning'))), '')
                        return matched_item.get('name', '')
                    elif typeData == 15:
                        matched_item = next((item for item in columns_text if int(item['id']) == typeData), '')
                        return matched_item.get('name', '')
            else:
                if 'ColumnName' in column:
                    return column.get('ColumnName')
        return ''

    def get_internal_key(column, type_name):
        if 'ColumnType' in column:
            ColumnType = column.get('ColumnType')
            if ColumnType == 'Selected':
                return f"{type_name}:{column.get('typeData')}:{column.get('variantPlaning')}"
        return ''

    def get_index_column_layout(columns_layout, simple_key, idx):
        col_index = 0
        for i, col in enumerate(columns_layout):
            if col['type'] == simple_key:
                if idx == col_index:
                    return i
                else:
                    col_index += 1
        return 0

    # def get_calced_formula_link_index_column_layout_old(columns_layout, simple_key, formula_link):
    #     FormulaLink = formula_link
    #     if 'total' in FormulaLink and FormulaLink.get('total') > 0:
    #         col_index = [0] * FormulaLink.get('total')
    #         for j in range(1,FormulaLink.get('total')+1):
    #             if f'col{j}' in FormulaLink:
    #                 FormulaLink[f'col{j}']['index'] = get_index_column_layout(
    #                     columns_layout,
    #                     simple_key,
    #                     FormulaLink.get(f'col{j}').get('index'))
    #     return FormulaLink
    def get_calced_formula_link_index_column_layout(formula_link):
        FormulaLink = formula_link
        if 'total' in FormulaLink and FormulaLink.get('total') > 0:
            col_index = [0] * FormulaLink.get('total')
            for j in range(1, FormulaLink.get('total') + 1):
                if f'col{j}' in FormulaLink:
                    FormulaLink[f'col{j}']['index'] = 0
        return FormulaLink

    def get_count_merge_group(columns, column):
        count = 0
        if column.get('IsStartGroup', False):
            MergeCount = column.get('MergeCount', 0)
            if MergeCount:
                return MergeCount
            else:
                NeedGroupName = column.get('GroupName')
                if NeedGroupName:
                    for col in columns:
                        if 'GroupName' in col and col.get('GroupName') == NeedGroupName:
                            count += 1
        return count

    template_name = g_report_template_name
    path_template = str(Path(uf.main_folder) / Path(uf.file_folder) / Path(template_name))

    # Получаем рабочую книгу из шаблона
    wb = openpyxl.load_workbook(path_template)

    # Создаём буфер для наполнения
    buffer = io.BytesIO()

    # ==================================================================================================================
    ws_tech = wb['tech']
    if ws_tech:
        G_STYLE_FONT_SIMPLE = extract_cell_styles(ws_tech['B2'])
        G_STYLE_FONT_FORMULA = extract_cell_styles(ws_tech['B3'])
        G_STYLE_FONT_TITLE_LVL0 = extract_cell_styles(ws_tech['B4'])
        G_STYLE_FONT_TITLE_LVL1 = extract_cell_styles(ws_tech['B5'])
        # ===========================================================
        G_STYLE_FULL_BORDER = extract_fast_border(ws_tech['B7'])
        G_STYLE_FULL_BORDER2 = NamedStyle(name="user_custom_border")
        G_STYLE_FULL_BORDER2.border = G_STYLE_FULL_BORDER
        G_STYLE_FULL_BORDER = True
        if "user_custom_border" not in wb.named_styles:
            wb.add_named_style(G_STYLE_FULL_BORDER2)
        # ===========================================================
        G_STYLE_FORMAT_PERCENTAGE = ws_tech['B8'].number_format
        G_STYLE_FONT_TITLE_LVL1_GROUP1 = extract_cell_styles(ws_tech['B9'])
        G_STYLE_FONT_TITLE_LVL2_GROUP1 = extract_cell_styles(ws_tech['B10'])
        G_STYLE_FONT_CHECKER1 = extract_cell_styles(ws_tech['B11'])

        G_STYLE_RULE_GREEN_RED_DASH = []
        for cf in ws_tech.conditional_formatting:
            # RULES: GREEN, RED, DASH
            if 'B6' in cf.cells:
                for rule in cf.rules:
                    G_STYLE_RULE_GREEN_RED_DASH.append(rule)
            # RULE: DASH
            if 'B12' in cf.cells:
                for rule in cf.rules:
                    G_STYLE_RULE_DASH_FOR_ZERO = rule
        G_STYLE_CELL_INPUT = extract_cell_styles(ws_tech['B13'])
    # ==================================================================================================================
    columns = []
    if generating_type.get('NeedGeneratingReport', False):
        for i, column in enumerate(src_columns):
            loc_column = copy.deepcopy(column)
            loc_column['ColumnType'] = 'Selected'
            loc_column['SrcKey'] = i
            columns.append(loc_column)
            columns.append({
                'ColumnType': 'PercentOfOutput',
                'ColumnName': '% выхода',
                'IsStartGroup': True,
                'MergeCount': 1,
                'GroupName': '% выхода',
                'FormulaLink': {
                    'Formula': '=ОКРУГЛ({}/{}*100;6)',
                    'total': 2,
                    'col1': {
                        'Letter': '',
                        'index': 0
                    },
                    'col2': {
                        'Letter': '',
                        'index': 0
                    }
                }
                # 'ptr': columns[1]
            })
        if len_src_columns > 1:
            columns.append({
                'ColumnType': 'SecondMinusFirst',
                'ColumnName': '+ / -',
                'IsStartGroup': True,
                'GroupName': 'Отклонение',
                'FormulaLink': {
                    'Formula': '=ЕСЛИОШИБКА({}-{};"-")',
                    'total': 2,
                    'col1': {
                        'Letter': '',
                        'index': 1
                    },
                    'col2': {
                        'Letter': '',
                        'index': 0
                    }
                }
                # 'ptr': columns[0]
            })
            columns.append({
                'ColumnType': 'PercentOfComplete',
                'ColumnName': '% вып.',
                'IsStartGroup': False,
                'GroupName': 'Отклонение',
                'FormulaLink': {
                    'Formula': '=ЕСЛИОШИБКА({}/{};"-")',
                    'total': 2,
                    'col1': {
                        'Letter': '',
                        'index': 1
                    },
                    'col2': {
                        'Letter': '',
                        'index': 0
                    }
                }
                # 'ptr': columns[0]
            })
        for column in columns:
            if column.get('ColumnType', '') != 'Selected':
                column['SrcKey'] = -1
        # ==============================================================================================================
        # composite_keys_do_pj = [(row.DO, row.pj) for row in settings]
        # bs_calc_mapping = get_data_from_query(
        #     """
        #                 SELECT "DO", pj, bs, calc FROM tab_bs_calc_map_d816_4 WHERE ("DO",pj) IN :composite_keys
        #             """,
        #     {"composite_keys": tuple(composite_keys_do_pj)})
        # ==============================================================================================================
        temp_data = {}
        for col in columns:
            if 'typeData' in col:
                type_id = int(col['typeData'])

                if type_id == 1:
                    if type_id not in temp_data:
                        temp_data[type_id] = []

                    temp_data[type_id].append({'variant_planing': col.get('variantPlaning')})

                else:
                    temp_data[type_id] = {'simple': True}

        columns_collect = [{key: val} for key, val in temp_data.items()]

        columns_text = []
        # for collect in columns_collect:
        for item in columns_collect:
            for type_id, collect in item.items():
                if type_id == 1:
                    variants = [item['variant_planing'] for item in collect]

                    query_res = uf.get_data_from_query(
                        'SELECT id, name FROM tab_view_var_plan_d816_4 WHERE id IN :variant_ids',
                        {'variant_ids': tuple(variants)})
                    for row in query_res:
                        columns_text.append({'typeData': type_id, 'id': row.id, 'name': row.name})

                elif type_id == 15:
                    query_res = uf.get_data_from_query(
                        'SELECT id, name FROM tab_view_io_bcblm0003_d816_4 WHERE id = :type_id',
                        {'type_id': type_id})
                    for row in query_res:
                        columns_text.append({'typeData': type_id, 'id': row.id, 'name': row.name})
        # ==============================================================================================================
        count_columns = len(columns)

        columns_layout = []
        for index_column, column in enumerate(columns):
            ColumnType = column.get('ColumnType')
            if ColumnType == 'Selected':
                simple_key = f"year{column.get('dateRange')[0][-4:]}"
                columns_layout.append({
                    'Letter': '',
                    'ColumnType': ColumnType,
                    'SrcKey': column.get('SrcKey'),
                    'type': simple_key,
                    'data_type_col': index_column,
                    'IsNeedMerge': True if index_column == 0 else False,
                    'MergeCount': count_columns,
                    'col_name': get_txt_col(column),
                    'internal_key': get_internal_key(column, simple_key),
                    'internal_key_xlsx': get_internal_key(column, simple_key),
                    'period': column.get('dateRange')[0][-4:]
                })
            elif ColumnType in ('SecondMinusFirst', 'PercentOfComplete', 'PercentOfOutput'):
                # if 'ptr' in column:
                #     col = column.get('ptr')
                #     simple_key = f"year{col.get('dateRange')[0][-4:]}"
                FormulaLink = copy.deepcopy(column.get('FormulaLink'))
                columns_layout.append({
                    'Letter': '',
                    'ColumnType': ColumnType,
                    'SrcKey': column.get('SrcKey'),
                    'type': 'year',
                    'data_type_col': index_column,
                    'IsNeedMerge': column.get('IsStartGroup', False),
                    'MergeCount': get_count_merge_group(columns, column),
                    'col_name': get_txt_col(column),
                    'GroupName': column.get('GroupName', ''),
                    'internal_key': '',  # get_internal_key(column, simple_key),
                    'internal_key_xls': '',
                    'FormulaLink': FormulaLink,
                    'period': ''
                })

        for q in range(1, 5):
            for index_column, column in enumerate(columns):
                ColumnType = column.get('ColumnType')
                if ColumnType == 'Selected':
                    simple_key = f"Q{q}"
                    columns_layout.append({
                        'Letter': '',
                        'ColumnType': ColumnType,
                        'SrcKey': column.get('SrcKey'),
                        'type': simple_key,
                        'data_type_col': index_column,
                        'IsNeedMerge': True if index_column == 0 else False,
                        'MergeCount': count_columns,
                        'col_name': get_txt_col(column),
                        'internal_key': get_internal_key(column, f"year{column.get('dateRange')[0][-4:]}:{simple_key}"),
                        'internal_key_xls': get_internal_key(column, f"year{column.get('dateRange')[0][-4:]}:{simple_key}"),
                        'period': column.get('dateRange')[0][-4:]
                    })
                elif ColumnType in ('SecondMinusFirst', 'PercentOfComplete', 'PercentOfOutput'):
                    # if 'ptr' in column:
                    #     col = column.get('ptr')
                    simple_key = f"Q{q}"
                    FormulaLink = copy.deepcopy(column.get('FormulaLink'))
                    columns_layout.append({
                        'Letter': '',
                        'ColumnType': ColumnType,
                        'SrcKey': column.get('SrcKey'),
                        'type': simple_key,
                        'data_type_col': index_column,
                        'IsNeedMerge': column.get('IsStartGroup', False),
                        'MergeCount': get_count_merge_group(columns, column),
                        'col_name': get_txt_col(column),
                        'GroupName': column.get('GroupName', ''),
                        'internal_key': '',  # get_internal_key(column, simple_key),
                        'internal_key_xls': '',
                        'FormulaLink': FormulaLink,
                        'period': ''
                    })
            for m in range(1, 4):
                for index_column, column in enumerate(columns):
                    ColumnType = column.get('ColumnType')
                    if ColumnType == 'Selected':
                        simple_key = f"M{q}_{m}"
                        calmonth = (q - 1) * 3 + m
                        columns_layout.append({
                            'Letter': '',
                            'ColumnType': ColumnType,
                            'SrcKey': column.get('SrcKey'),
                            'type': simple_key,
                            'data_type_col': index_column,
                            'IsNeedMerge': True if index_column == 0 else False,
                            'MergeCount': count_columns,
                            'col_name': get_txt_col(column),
                            'calmonth': calmonth,
                            'internal_key': get_internal_key(column,f"year{column.get('dateRange')[0][-4:]}:{simple_key}"),
                            'internal_key_xls': get_internal_key(column,f"year{column.get('dateRange')[0][-4:]}:M{q}_{calmonth}"),
                            'period': column.get('dateRange')[0][-4:]
                        })
                    elif ColumnType in ('SecondMinusFirst', 'PercentOfComplete', 'PercentOfOutput'):
                        # if 'ptr' in column:
                        #     col = column.get('ptr')
                        simple_key = f"M{q}_{m}"
                        FormulaLink = copy.deepcopy(column.get('FormulaLink'))
                        columns_layout.append({
                            'Letter': '',
                            'ColumnType': ColumnType,
                            'SrcKey': column.get('SrcKey'),
                            'type': simple_key,
                            'data_type_col': index_column,
                            'IsNeedMerge': column.get('IsStartGroup', False),
                            'MergeCount': get_count_merge_group(columns, column),
                            'col_name': get_txt_col(column),
                            'GroupName': column.get('GroupName', ''),
                            'internal_key': '',  # get_internal_key(column, simple_key),
                            'internal_key_xls': '',
                            'FormulaLink': FormulaLink,
                            'period': ''
                        })
        #  Добавляем в макет данные под годовые столбцы трёхлетки (+1 и +2 года, т.е. по 2 столбца на каждый вариант)
        for y in range(1, 3):
            for index_column, column in enumerate(columns):
                ColumnType = column.get('ColumnType')
                if ColumnType == 'Selected' and column.get('IsByear', False):
                    year = int(column.get('dateRange')[0][-4:]) + y
                    simple_key = f"Byear{year}"
                    columns_layout.append({
                        'Letter': '',
                        'ColumnType': 'Byear',
                        'SrcKey': column.get('SrcKey'),
                        'SrcKeyByear': len_src_columns * y + int(column.get('SrcKey')),
                        'type': simple_key,
                        'data_type_col': index_column,
                        'IsNeedMerge': False,
                        'MergeCount': 0,
                        'col_name': get_txt_col(column),
                        'internal_key': get_internal_key(column, f"year{year}:{simple_key}"),
                        'internal_key_xls': get_internal_key(column, f"year{year}:{simple_key}"),
                        'period': year
                    })
        date_format = "%d.%m.%Y"

        # len_col = len(columns)
        # for y in range(1, 3):
        #     for i in range(len_col):
        #         ColumnType = columns[i].get('ColumnType', '')
        #         if ColumnType == 'Selected':
        #             col = copy.deepcopy(columns[i])
        #             col['ColumnType'] = 'Byear'
        #             col['SrcKey'] += len_src_columns * y
        #
        #             new_dates = []
        #             for date_str in col["dateRange"]:
        #                 date_obj = datetime.strptime(date_str, date_format)
        #                 new_date_obj = date_obj.replace(year=date_obj.year + y)
        #
        #                 new_dates.append(new_date_obj.strftime(date_format))
        #
        #             col["dateRange"] = new_dates
        #
        #             columns.append(col)

    # ==================================================================================================================
    # Удаление неиспользуемых листов
    # if download_type == 'simple':
    #     sheets_to_keep = set()
    #
    #     for factory_id in selected_factories:
    #         range_name = f"_BS{factory_id}"
    #         defined_name = wb.defined_names.get(range_name)
    #
    #         if not defined_name:
    #             continue
    #
    #         try:
    #             for sheet_title, cell_range in defined_name.destinations:
    #                 if sheet_title in wb.sheetnames:
    #                     sheets_to_keep.add(sheet_title)
    #                 break
    #         except Exception as e:
    #             pass
    #
    #     if not sheets_to_keep:
    #         return False
    #
    #     for sheet_name in wb.sheetnames:
    #         if sheet_name not in sheets_to_keep and sheet_name != 'tech':
    #             wb.remove(wb[sheet_name])
    # ==================================================================================================================

    # [GENERATING] Инициализация работы для динамического построения отчёта
    def init_some_data_generating_report(type_sheet, sheet_ids, sheet_all, named_rng_names):
        # _SET_ROW:
        #
        nonlocal storage_sheet

        # for sheet_id in sheet_ids:
        for sheet_id in sheet_all:
            try:
                def_range_set_row = wb.defined_names[f'{named_rng_names[0]}{sheet_id}']
                def_range_ik = wb.defined_names[f'{named_rng_names[1]}{sheet_id}']
            except Exception as e:
                continue

            sheet_name_set_row = ''
            set_row_left_col_index = -1
            set_row_right_col_index = -1
            FirstRowData = -1
            LastRowData = -1
            first_row = -1

            for sheet_name_rng_bs, cell_coordinates_rng_set_row in def_range_set_row.destinations:
                set_row_left_col_index, _, set_row_right_col_index, _ = range_boundaries(cell_coordinates_rng_set_row)
                sheet_name_set_row = sheet_name_rng_bs

            if sheet_name_set_row in wb.sheetnames:
                sheet = wb[sheet_name_set_row]
            else:
                continue

            if sheet_id not in sheet_ids:
                sheet.sheet_state = 'veryHidden'
                continue

            if type_sheet == 'type_summary_rep':
                # Для СВОДОВ считать процент выхода нет необходимости
                struct_columns = []
                for col in columns:
                    if col.get('ColumnType') != 'PercentOfOutput':
                        struct_columns.append(copy.deepcopy(col))

                sheet_columns_layout = []
                for col in columns_layout:
                    if col.get('ColumnType') != 'PercentOfOutput':
                        loc_col = copy.deepcopy(col)
                        if loc_col and loc_col.get('IsNeedMerge', False):
                            loc_col['MergeCount'] = get_count_merge_group(struct_columns,
                                                                          columns[col.get('data_type_col')])
                        sheet_columns_layout.append(loc_col)
            else:
                sheet_columns_layout = copy.deepcopy(columns_layout)
                struct_columns = copy.deepcopy(columns)

            len_columns = len(struct_columns)

            storage_sheet[sheet_name_set_row]['columns_layout'] = sheet_columns_layout
            storage_sheet[sheet_name_set_row]['struct_columns'] = struct_columns

            storage_sheet[sheet_name_set_row]['columns_settings'] = get_common_column_settings(set_row_left_col_index)

            storage_sheet[sheet_name_set_row]['SET_ROW_LIST'] = {}
            storage_sheet[sheet_name_set_row]['SET_ROW_IDX'] = {}
            storage_sheet[sheet_name_set_row]['SET_ROW_KEY'] = {}
            storage_sheet[sheet_name_set_row]['SET_ROW_CHECK'] = {}

            # ==========================================================================================================
            Byear = []
            storage_sheet[sheet_name_set_row]['Byear'] = Byear
            len_col = len(columns)
            for y in range(1, 3):
                for i in range(len_col):
                    ColumnType = columns[i].get('ColumnType', '')
                    if ColumnType == 'Selected' and columns[i].get('IsByear', False):
                        col = copy.deepcopy(columns[i])
                        col['ColumnType'] = 'Byear'
                        col['SrcKey'] += len_src_columns * y

                        new_dates = []
                        for date_str in col["dateRange"]:
                            date_obj = datetime.strptime(date_str, date_format)
                            new_date_obj = date_obj.replace(year=date_obj.year + y)

                            new_dates.append(new_date_obj.strftime(date_format))

                        col["dateRange"] = new_dates

                        Byear.append(col)
            # ==========================================================================================================

            for key, idx_col in storage_sheet[sheet_name_set_row]['columns_settings'].items():
                if key not in storage_sheet[sheet_name_set_row]['SET_ROW_LIST']:
                    storage_sheet[sheet_name_set_row]['SET_ROW_LIST'][key] = []
                if key not in storage_sheet[sheet_name_set_row]['SET_ROW_KEY']:
                    storage_sheet[sheet_name_set_row]['SET_ROW_KEY'][key] = {}
                if key not in storage_sheet[sheet_name_set_row]['SET_ROW_CHECK']:
                    storage_sheet[sheet_name_set_row]['SET_ROW_CHECK'][key] = {}

            for sheet_name_rng_ik, cell_coordinates_rng_ik in def_range_ik.destinations:
                _, FirstRowData, _, _ = range_boundaries(cell_coordinates_rng_ik)

            if FirstRowData == -1:
                continue
            else:
                first_row = FirstRowData - 3  # всего строк для заголовков 3
                FirstRowData += 1

            if sheet and set_row_right_col_index != -1:
                dict_bs = []
                loc_index_offset = set_row_right_col_index + 1 + offset_ind_col
                loc_bs_calc_mapping = []
                last_row = sheet.max_row

                for i in range(1, last_row + 1):
                    # Сбор настроек в виде мэпинга по строкам с каждого листа
                    # из именованного диапазона, содержащего, *_SET_ROW*
                    if i >= FirstRowData:
                        for key, idx_col in storage_sheet[sheet_name_set_row]['columns_settings'].items():
                            cell_value = sheet.cell(row=i, column=idx_col).value
                            if cell_value is not None:
                                cell_value = str(cell_value)
                                storage_sheet[sheet_name_set_row]['SET_ROW_LIST'][key].append({
                                    i: cell_value
                                })

                                if i not in storage_sheet[sheet_name_set_row]['SET_ROW_IDX']:
                                    storage_sheet[sheet_name_set_row]['SET_ROW_IDX'][i] = {}
                                storage_sheet[sheet_name_set_row]['SET_ROW_IDX'][i][key] = {
                                    'column': idx_col,
                                    'value': cell_value
                                }

                                storage_sheet[sheet_name_set_row]['SET_ROW_KEY'][key][cell_value] = i

                                # Нет необходимости проверять на большее значение, т.к. основной цикл идёт по строкам сверху вниз
                                storage_sheet[sheet_name_set_row]['LastRowData'] = i
                                if storage_sheet[sheet_name_set_row].get('FirstRowWithBS', '') == '':
                                    storage_sheet[sheet_name_set_row]['FirstRowWithBS'] = i

                # Способ хранения литералов и индексов для столбцов макета листа:
                # {номер группы}:{индекс в группе}
                #
                # Пример:
                # группа 0 - {0:0} {0:1} {0:2} {0:3} {0:4}
                # группа 1 - {1:0} {1:1} {1:2} {1:3} {1:4}
                # группа 2 - {2:0} {2:1} {2:2} {2:3} {2:4}
                # ...
                #
                # Итоговый объём может выглядеть так: {0-16:0-4} 17 групп из 5 элементов
                for idx, col in enumerate(sheet_columns_layout):
                    col_num = idx + loc_index_offset
                    col_letter = get_column_letter(col_num)

                    unique_key_column = f'{idx // len_columns}:{col.get("data_type_col")}'
                    storage_sheet[sheet_name_set_row]['unique_key_column'][unique_key_column] = {
                        'index': col_num,
                        'Letter': col_letter
                    }

                    src_key_column = f'{idx // len_columns}:{col.get("SrcKey")}'
                    storage_sheet[sheet_name_set_row]['src_key_column'][src_key_column] = {
                        'index': col_num,
                        'Letter': col_letter
                    }

                    storage_sheet[sheet_name_set_row]['col'][idx] = {
                        'index': col_num,
                        'Letter': col_letter
                    }

                for idx, col in enumerate(sheet_columns_layout):
                    FormulaLink = col.get('FormulaLink', '')
                    if FormulaLink:
                        total = FormulaLink.get('total')
                        for i in range(1, total + 1):
                            formula_col = FormulaLink.get(f'col{i}', '')
                            if formula_col:
                                formula_col['Letter'] = storage_sheet[sheet_name_set_row]['src_key_column'] \
                                    [f'{idx // len_columns}:{formula_col["index"]}'].get('Letter')

    # [GENERATING] Подготовка и заполнение листов с динамическим построением отчёта
    def prepare_and_fill_data_generating_report(type_sheet, sheet_ids, named_rng_names):
        nonlocal storage_sheet

        for sheet_id in sheet_ids:
            try:
                def_range_bs = wb.defined_names[f'{named_rng_names[0]}{sheet_id}']
                def_range_ik = wb.defined_names[f'{named_rng_names[1]}{sheet_id}']
            except Exception as e:
                continue

            sheet_name_set_row = ''
            set_row_left_col_index = -1
            set_row_right_col_index = -1
            FirstRowData = -1
            LastRowData = -1
            first_row = -1

            for sheet_name_rng_bs, cell_coordinates_rng_set_row in def_range_bs.destinations:
                set_row_left_col_index, _, set_row_right_col_index, _ = range_boundaries(cell_coordinates_rng_set_row)
                sheet_name_set_row = sheet_name_rng_bs

            if sheet_name_set_row in wb.sheetnames:
                sheet = wb[sheet_name_set_row]
            else:
                continue

            for sheet_name_rng_ik, cell_coordinates_rng_ik in def_range_ik.destinations:
                _, FirstRowData, _, _ = range_boundaries(cell_coordinates_rng_ik)

            if FirstRowData == -1:
                continue
            else:
                first_row = FirstRowData - 3  # всего строк для заголовков 3
                FirstRowData += 1

            if sheet:
                sheet.sheet_state = 'visible'

            if sheet and set_row_right_col_index != -1:
                nr_sheet_id = get_sheet_id_from_sheet(sheet)

                sheet_columns_layout = storage_sheet[sheet_name_set_row]['columns_layout']
                struct_columns = storage_sheet[sheet_name_set_row]['struct_columns']
                key_bs = storage_sheet[sheet_name_set_row]['SET_ROW_LIST'][EnumColumnSettings.KEY_BS]
                key_input = storage_sheet[sheet_name_set_row]['SET_ROW_LIST'][EnumColumnSettings.KEY_INPUT]
                LastRowData = storage_sheet[sheet_name_set_row]['LastRowData']
                SetRowIdx = storage_sheet[sheet_name_set_row]['SET_ROW_IDX']
                Byear = storage_sheet[sheet_name_set_row]['Byear']

                dict_bs = []
                dict_key_input = []
                loc_index_offset = set_row_right_col_index + 1 + offset_ind_col
                loc_bs_calc_mapping = []
                last_row = sheet.max_row

                for item_dict in key_bs:
                    value = list(item_dict.values())[0]
                    if value is not None and str(value).isdigit():
                        dict_bs.append(value)
                for item_dict in key_input:
                    for row, key_value in item_dict.items():
                        if key_value is not None:
                            dict_key_input.append(key_value)

                if G_STYLE_FULL_BORDER:
                    _min_col = set_row_right_col_index + 1 + offset_ind_col
                    _max_col = set_row_right_col_index + 1 + offset_ind_col + len(sheet_columns_layout) - 1

                    cells_dict = sheet._cells

                    for row_idx in range(first_row, LastRowData + 1):
                        for col_idx in range(_min_col, _max_col + 1):
                            coords = (row_idx, col_idx)

                            if coords not in cells_dict:
                                cells_dict[coords] = Cell(sheet, row=row_idx, column=col_idx)

                            cells_dict[coords].style = "user_custom_border"

                sheet.row_dimensions[first_row + 0].height = 20  # column title
                sheet.row_dimensions[first_row + 1].height = 40  # column name
                sheet.row_dimensions[first_row + 2].height = 20  # column helper description
                sheet.row_dimensions[first_row + 3].height = 20  # column internal_key

                multi_range_rules = MultiCellRange()
                multi_range_rule_dash_for_zero = MultiCellRange()
                for idx, col in enumerate(sheet_columns_layout):
                    col_num = idx + loc_index_offset
                    col_letter = get_column_letter(col_num)
                    col['Letter'] = col_letter
                    ColumnType = col.get('ColumnType')

                    # Уровни заголовков:
                    #           first_row
                    # 1 lvl (first_row + 1)
                    # 2 lvl (first_row + 2)
                    # 3 lvl (first_row + 3)
                    #           key_row
                    # data_row1
                    # data_row2
                    # data_row3
                    # ...

                    if ColumnType == 'Selected' or ColumnType == 'Byear':
                        # Заголовок уровня [1-2] (объединённый)
                        cell = sheet.cell(row=first_row + 1, column=col_num)
                        sheet.merge_cells(start_row=first_row + 1, start_column=col_num, end_row=first_row + 2,
                                          end_column=col_num)
                        set_value_cell(cell, col.get('col_name'), EnumCellType.TITLE_LVL0)

                        sheet.column_dimensions[col_letter].width = 25  # 22 #16.29
                    else:
                        # Заголовок уровня 2
                        cell = sheet.cell(row=first_row + 2, column=col_num)
                        if G_STYLE_RULE_GREEN_RED_DASH and ColumnType in ('SecondMinusFirst', 'PercentOfComplete'):
                            current_range = f'${col_letter}${FirstRowData}:${col_letter}${LastRowData}'
                            multi_range_rules.add(current_range)
                        if G_STYLE_RULE_DASH_FOR_ZERO and ColumnType == 'PercentOfOutput':
                            current_range = f'${col_letter}${FirstRowData}:${col_letter}${LastRowData}'
                            multi_range_rule_dash_for_zero.add(current_range)

                        if ColumnType != 'Selected':
                            set_value_cell(cell, col.get('col_name'), EnumCellType.TITLE_GROUP1_LVL2)
                        else:
                            set_value_cell(cell, col.get('col_name'), EnumCellType.TITLE_LVL0)

                        sheet.column_dimensions[col_letter].width = 15

                    if ColumnType == 'PercentOfOutput':
                        sheet.column_dimensions.group(start=col_letter, end=col_letter, hidden=True)

                    if ColumnType == 'Selected':
                        # Добавляем в список правил ячейки с проверками
                        for row in storage_sheet[sheet_name_set_row]['SET_ROW_LIST'][EnumColumnSettings.ROW_CHECK]:
                            row_index = next(iter(row))
                            current_range = f'${col_letter}${row_index}'
                            multi_range_rules.add(current_range)

                    # Заголовок уровня 3 (внутренние ключи)
                    # set_value_cell(sheet.cell(row=first_row + 3, column=col_num), col.get('internal_key'))
                    set_value_cell(sheet.cell(row=first_row + 3, column=col_num), col.get('internal_key_xls'))

                    if 'IsNeedMerge' in col and col['IsNeedMerge']:
                        MergeCount = col.get('MergeCount', 0)
                        if MergeCount > 1:
                            if ColumnType == 'Selected':
                                q = 0
                                # Заголовок уровня 0
                                ### sheet.merge_cells(start_row=first_row, start_column=col_num, end_row=first_row,
                                ###                   end_column=col_num + col['MergeCount'] - 1)
                            else:
                                # Заголовок уровня 1 (на строку ниже)
                                sheet.merge_cells(start_row=first_row + 1, start_column=col_num, end_row=first_row + 1,
                                                  end_column=col_num + MergeCount - 1)
                        elif MergeCount == 1:
                            sheet.merge_cells(start_row=first_row + 1, start_column=col_num,
                                              end_row=first_row + 2,
                                              end_column=col_num + MergeCount - 1)

                        if ColumnType == 'Selected':
                            q = 0
                            # cell = sheet.cell(row=first_row, column=col_num)
                            # if 'year' in col['type']:
                            #     set_value_cell(cell,col.get("period"), EnumColumnType.TITLE_LVL1)
                            # elif 'Q' in col['type']:
                            #     q_number = col['type'][1]
                            #     set_value_cell(cell,f'{col.get("period")} год {q_number} квартал', EnumColumnType.TITLE_LVL1)
                            # elif 'M' in col['type']:
                            #     set_value_cell(cell,f'{col.get("period")} год {month[col.get("calmonth")]}', EnumColumnType.TITLE_LVL1)
                        else:
                            GroupName = col.get('GroupName', '')
                            if GroupName:
                                cell = sheet.cell(row=first_row + 1, column=col_num)
                                set_value_cell(cell, GroupName, EnumCellType.TITLE_GROUP1_LVL1)

                    if ColumnType == 'Selected' or ColumnType == 'Byear':
                        cell = sheet.cell(row=first_row, column=col_num)
                        if 'year' in col['type']:
                            set_value_cell(cell, col.get("period"), EnumCellType.TITLE_LVL1)
                        elif 'Q' in col['type']:
                            q_number = col['type'][1]
                            set_value_cell(cell, f'{col.get("period")} год {q_number} квартал',
                                           EnumCellType.TITLE_LVL1)
                        elif 'M' in col['type']:
                            set_value_cell(cell, f'{col.get("period")} год {month[col.get("calmonth")]}',
                                           EnumCellType.TITLE_LVL1)

                if G_STYLE_RULE_GREEN_RED_DASH:
                    for rule in G_STYLE_RULE_GREEN_RED_DASH:
                        sheet.conditional_formatting.add(str(multi_range_rules), rule)
                if G_STYLE_RULE_DASH_FOR_ZERO:
                    sheet.conditional_formatting.add(str(multi_range_rule_dash_for_zero), G_STYLE_RULE_DASH_FOR_ZERO)

                query_res_for_column = []
                do = 0
                pj = 0
                if type_sheet == 'type_factory':
                    loc_settings = next((row for row in settings if row.id == sheet_id), None)
                    if loc_settings:
                        do = loc_settings.DO
                        pj = loc_settings.pj
                elif type_sheet == 'type_summary_rep':
                    pass
                    # for index_column, column in enumerate(struct_columns, start=1):
                    #     if column.get('ColumnType') == 'Selected':
                    #         query_res_for_column.append(0)

                for index_column, column in enumerate(struct_columns, start=1):
                    if column.get('ColumnType') == 'Selected':
                        year = column.get('dateRange')[0][-4:]
                        ver_plan = column.get('versionPlaning', 0)
                        var_plan = column.get('variantPlaning', 0)
                        data_type = int(column.get('typeData'))
                        query_res_for_column.append(
                            get_row_list_msb_zuv_d816_4(year,
                                                        ver_plan,
                                                        var_plan,
                                                        dict_bs,
                                                        dict_key_input,
                                                        do,
                                                        pj,
                                                        data_type,
                                                        nr_sheet_id))
                if Byear:
                    for column in Byear:
                        year = column.get('dateRange')[0][-4:]
                        ver_plan = column.get('versionPlaning', 0)
                        var_plan = column.get('variantPlaning', 0)
                        data_type = int(column.get('typeData'))
                        query_res_for_column.append(
                            get_row_list_msb_zuv_d816_4(year,
                                                        ver_plan,
                                                        var_plan,
                                                        dict_bs,
                                                        dict_key_input,
                                                        do,
                                                        pj,
                                                        data_type,
                                                        nr_sheet_id,
                                                        True))

                # основная функция для заполнения таблицы подготовленными данными
                fill_flat_data(type_sheet,
                               offset_ind_col,
                               sheet,
                               set_row_right_col_index,
                               FirstRowData,
                               LastRowData,
                               sheet_columns_layout,
                               query_res_for_column,
                               download_type,
                               struct_columns,
                               storage_sheet)

    # [STATIC] Инициализация работы для динамического построения отчёта
    def init_some_data_static_report(selected_report, named_rng_names):
        nonlocal storage_sheet

        # ===============================================================================================================
        # SET_ROW
        named_rng_data = get_data_from_named_range_name(wb, named_rng_names.get('SET_ROW'))
        if named_rng_data.get('Exec', False) == False:
            return
        sheet_name_set_row = named_rng_data.get('sheet_name')
        set_row_left_col_index = named_rng_data.get('min_col')
        set_row_right_col_index = named_rng_data.get('max_col')
        sheet = named_rng_data.get('sheet')
        # ===============================================================================================================
        # INTERNAL_KEY
        named_rng_data = get_data_from_named_range_name(wb, named_rng_names.get('INTERNAL_KEY'))
        if named_rng_data.get('Exec', False) == False:
            return
        # ===============================================================================================================
        sheet.sheet_state = 'visible'
        required_conditions = set()

        FirstRowData = named_rng_data.get('min_row')
        first_row = FirstRowData - 3  # всего строк для заголовков 3
        FirstRowData += 1

        storage_sheet[sheet_name_set_row]['columns_layout'] = {
            '1': get_column_letter(set_row_right_col_index + 3),
            '15': get_column_letter(set_row_right_col_index + 6)
        }
        collect_struct_columns = {
            '1': {
                'filled': False,
                'versionPlaning': 0,
                'variantPlaning': 0,
                'dateRange': [],
            },
            '15': {
                'filled': False,
                'dateRange': [],
            }
        }
        for col in src_columns:
            type_data = col.get('typeData')
            if type_data == '1' and collect_struct_columns[type_data].get('filled', False) == False:
                collect_struct_columns[type_data]['filled'] = True
                collect_struct_columns[type_data]['versionPlaning'] = col.get('versionPlaning')
                collect_struct_columns[type_data]['variantPlaning'] = col.get('variantPlaning')
                collect_struct_columns[type_data]['dateRange'] = col.get('dateRange')
            if type_data == '15':
                collect_struct_columns[type_data]['filled'] = True
                collect_struct_columns[type_data]['dateRange'] = col.get('dateRange')
        storage_sheet[sheet_name_set_row]['struct_columns'] = collect_struct_columns

        storage_sheet[sheet_name_set_row]['columns_settings'] = {
            EnumColumnSettings.FORMULA_MONTH: set_row_left_col_index + 0,
        }

        storage_sheet[sheet_name_set_row]['SET_ROW_LIST'] = {}
        storage_sheet[sheet_name_set_row]['SET_ROW_IDX'] = {}
        storage_sheet[sheet_name_set_row]['required_conditions'] = required_conditions
        storage_sheet[sheet_name_set_row]['collect_bs'] = []

        for key, idx_col in storage_sheet[sheet_name_set_row]['columns_settings'].items():
            if key not in storage_sheet[sheet_name_set_row]['SET_ROW_LIST']:
                storage_sheet[sheet_name_set_row]['SET_ROW_LIST'][key] = []

        loc_index_offset = set_row_right_col_index + 1 + offset_ind_col
        last_row = sheet.max_row

        for i in range(1, last_row + 1):
            if i >= FirstRowData:
                for key, idx_col in storage_sheet[sheet_name_set_row]['columns_settings'].items():
                    cell_value = sheet.cell(row=i, column=idx_col).value
                    if cell_value is not None:
                        cell_value = str(cell_value)
                        storage_sheet[sheet_name_set_row]['SET_ROW_LIST'][key].append({
                            i: cell_value
                        })

                        if i not in storage_sheet[sheet_name_set_row]['SET_ROW_IDX']:
                            storage_sheet[sheet_name_set_row]['SET_ROW_IDX'][i] = {}
                        storage_sheet[sheet_name_set_row]['SET_ROW_IDX'][i][key] = {
                            'column': idx_col,
                            'value': cell_value
                        }

                        # Нет необходимости проверять на большее значение, т.к. основной цикл идёт по строкам сверху вниз
                        storage_sheet[sheet_name_set_row]['LastRowData'] = i
                        if storage_sheet[sheet_name_set_row].get('FirstRowWithBS', '') == '':
                            storage_sheet[sheet_name_set_row]['FirstRowWithBS'] = i

                        if key == EnumColumnSettings.FORMULA_MONTH:
                            # ===============================================================================================
                            new_value = '0'
                            src_formula = ''
                            pattern = re.compile(get_hard_mirror_pattern())
                            formula = cell_value
                            for match in pattern.finditer(formula):
                                token_type = match.lastgroup
                                value = match.group()
                                new_value = '0'

                                # Округл - Имя функции формулы-xls
                                # 1000330001 - ключ статьи
                                # 1_1000330001 - ключ статьи со ссылкой на лист
                                if token_type == 'FUNC_OR_VAR':
                                    end_index = match.end()
                                    next_char = formula[end_index] if end_index < len(formula) else ''
                                    # Проверяем, если попало имя функции, то просто передаём его дальше без обработки
                                    if next_char == '(':
                                        new_value = f'{value}'
                                    elif re.match(r'^\$?[A-Za-z]{1,3}\$?\d+$', value):
                                        # Пропускаем обработку формул-xls (A1+B2)
                                        pass
                                    else:
                                        if '_' in value:
                                            index, bs = value.split('_')
                                            if index and bs:
                                                index_do = int(index)
                                                index = index_do - 1
                                                do = settings[index]._mapping['DO']
                                                pj = settings[index]._mapping['pj']
                                                required_conditions.add((index_do, do, pj, int(bs)))
                                        else:
                                            # Нет логики обработки без ссылочной статьи (статья без номера завода)
                                            pass
                                # экранирование для $
                                elif token_type == 'ESCAPED':
                                    new_value = value.replace('$', '')

                                # обычный элемент
                                else:
                                    new_value = value

                            src_formula = f'{src_formula}{new_value}'
                        # ===============================================================================================
        if required_conditions:
            # Запрос на получение цифр
            db = uf.get_db_connection()

            # Разделяем параметры на 3 изолированных списка
            params_list = list(required_conditions)
            index_do = [p[0] for p in params_list]
            do = [p[1] for p in params_list]
            pj = [p[2] for p in params_list]
            bs = [p[3] for p in params_list]

            query_sql = text(f"""
                SELECT
                    p.req_year AS year,
                    t.BCBLM0002::INT as var_planing, -- Вариант планирования
                    p.req_data_type AS data_type,
                    f.index_do AS index,
                    f.req_do AS do,
                    f.req_pj AS pj,
                    f.req_bs AS bs,
                    COALESCE(SUM(t.sum), 0) AS sum
                FROM
                unnest(
                        CAST(:index_do AS INTEGER[]),   -- [Набор] Индекс do
                        CAST(:do AS INTEGER[]),         -- [Набор] Завод
                        CAST(:pj AS INTEGER[]),         -- [Набор] Перерабатывающий комплекс (Поставщики ЖУВ)
                        CAST(:bs AS INTEGER[])          -- [Набор] Бюджетная статья
                            )AS f(index_do, req_do, req_pj, req_bs)
                CROSS JOIN 
                    (VALUES
                        ({collect_struct_columns['1']['dateRange'][0][-4:]},
                        {collect_struct_columns['1']['versionPlaning']},
                        {collect_struct_columns['1']['variantPlaning']},
                        1),
                        
                        ({collect_struct_columns['15']['dateRange'][0][-4:]},
                        0,
                        0,
                        15)
                    ) AS p(req_year, req_ver_planing, req_var_planing, req_data_type)
                
                LEFT JOIN 
                    tab_integ_get_preu_mirror_d816_4 AS t
                    ON  t.BCBIM0002::INT = f.req_do 
                    AND t.pj::INT        = f.req_pj 
                    AND t.bs::INT        = f.req_bs
                    AND t.calyear        = p.req_year
                    AND t.BCBLM0001::INT = p.req_ver_planing
                    AND t.BCBLM0002::INT = p.req_var_planing
                    AND t.data_type::INT = p.req_data_type
                    AND t.dbs = 0 
                    AND t.calmonth <> 0
                GROUP BY
                    p.req_year,
                    t.BCBLM0002,
                    p.req_data_type,
                    f.index_do,
                    f.req_do,
                    f.req_pj,
                    f.req_bs
                ORDER BY
                    year,
                    data_type,
                    index,
                    BCBLM0002,
                    pj,
                    bs
            """)
            result = db.execute(query_sql, {
                'index_do': index_do,
                'do': do,
                'pj': pj,
                'bs': bs,
            }).fetchall()
            if result:
                storage_sheet[sheet_name_set_row]['collect_bs'] = res = [row._asdict() for row in result]

    # [STATIC] Подготовка и заполнение листов с динамическим построением отчёта
    def prepare_and_fill_data_static_report(selected_report, named_rng_names):
        nonlocal storage_sheet

        # ==============================================================================================================
        # SET_ROW
        named_rng_data = get_data_from_named_range_name(wb, named_rng_names.get('SET_ROW'))
        if named_rng_data.get('Exec', False) == False:
            return
        sheet_name = named_rng_data.get('sheet_name')
        set_row_left_col_index = named_rng_data.get('min_col')
        set_row_right_col_index = named_rng_data.get('max_col')
        sheet = named_rng_data.get('sheet')
        # ==============================================================================================================
        # INTERNAL_KEY
        named_rng_data = get_data_from_named_range_name(wb, named_rng_names.get('INTERNAL_KEY'))
        if named_rng_data.get('Exec', False) == False:
            return
        # ==============================================================================================================
        FirstRowData = named_rng_data.get('min_row')
        first_row = FirstRowData - 3  # всего строк для заголовков 3
        FirstRowData += 1

        SetRowIdx = storage_sheet[sheet_name]['SET_ROW_IDX']
        columns_layout = storage_sheet[sheet_name]['columns_layout']
        collect_bs = storage_sheet[sheet_name]['collect_bs']
        collect_struct_columns = storage_sheet[sheet_name]['struct_columns']

        for i, v in SetRowIdx.items():
            for data_type, Letter in columns_layout.items():
                src_formula = v[EnumColumnSettings.FORMULA_MONTH]['value']
                src_formula = src_formula.replace('$', '')
                for elem in collect_bs:
                    if elem.get("data_type") == int(data_type) and \
                            elem.get("year") == int(collect_struct_columns[data_type]['dateRange'][0][-4:]):
                        src_formula = src_formula.replace(
                            f'{elem.get("index")}_{elem.get("bs")}', str(float(elem.get("sum"))))
                set_value_cell(sheet[f"{Letter}{i}"],
                               formula_translator.convert_russian_formula(src_formula))

    if generating_type.get('NeedGeneratingReport', False):
        # generating_report_settings = {
        # # ключ в "Таблица Типы отчётов" | index - индекс именованного диапазона

        reports_all_list = reports_all_list = [value.get('index') for key, value in generating_report_settings.items()]
        selected_index_report = [
            generating_report_settings[report_id]['index']
            for report_id in selected_reports
            if report_id in generating_report_settings
        ]

        init_some_data_generating_report('type_factory', selected_factories, factories_all,
                                         ['_SET_ROW', '_INTERNAL_KEY'])
        init_some_data_generating_report('type_summary_rep', selected_index_report, reports_all_list,
                                         ['_SUM_REP_SET_ROW', '_SUM_REP_INTERNAL_KEY'])

        prepare_and_fill_data_generating_report('type_factory', selected_factories, ['_SET_ROW', '_INTERNAL_KEY'])
        prepare_and_fill_data_generating_report('type_summary_rep', selected_index_report,
                                                ['_SUM_REP_SET_ROW', '_SUM_REP_INTERNAL_KEY'])

    if generating_type.get('NeedStaticReport', False):
        #FIXME Доработать получение настроек из настроечной таблицы, которая реализована через функцию:
        #FIXME get_sheets_settings()
        static_report_settings = {
            '22': {  # Отчёт КПД ДО
                'SET_ROW': '_STATIC_REP_SET_ROW1',
                'INTERNAL_KEY': '_STATIC_REP_INTERNAL_KEY1',
            }
        }
        for sel_rep in selected_reports:
            if static_report_settings.get(sel_rep, False):
                init_some_data_static_report(sel_rep, static_report_settings[sel_rep])
        for sel_rep in selected_reports:
            if static_report_settings.get(sel_rep, False):
                prepare_and_fill_data_static_report(sel_rep, static_report_settings[sel_rep])

    # ===================================================================================================================

    # Сохраняем подготовленные данные из шаблона
    wb.save(buffer)
    # Откатываем курсор в самое начало
    buffer.seek(0)
    # имя файла
    filename = template_name

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============================================================================================
def fill_flat_data(
        type_sheet,
        offset_ind_col,
        sheet,
        bs_col_index,
        FirstRowData,
        LastRowData,
        sheet_columns_layout,
        query_res_for_column,
        download_type,
        struct_columns,
        loc_storage_sheet):
    len_columns = len(struct_columns)

    def get_col_letter_by_type(col_type, data_type_col):
        for idx, col in enumerate(sheet_columns_layout):
            if col["type"] == col_type and col["data_type_col"] == data_type_col:
                return col.get('Letter')
        return None

    def set_calc_cell_val(cell, query_cell, col_letter, query_res, col, idx, ColumnSettings=EnumColumnSettings.FORMULA_MONTH,
                          ColumnCellType: EnumCellType = EnumCellType.INPUT):

        def get_sheetname_from_sheetkey(sheet_key):
            sheet_name = ''
            named_rng_ik_name = f'_INTERNAL_KEY{sheet_key}'
            if named_rng_ik_name in wb.defined_names:
                nr_rng = wb.defined_names[named_rng_ik_name]
                destinations = list(nr_rng.destinations)
                if destinations:
                    sheet_name, cell_coord = destinations[0]
            return sheet_name

        # 1_10030001 -> Лист1!AB4 или 1_10030001 -> 4 (эта логика нереализована[OnlyRow=True])
        def get_cell_addr_by_linked_bs(p, OnlyRow=False):
            sheet_key, bs = p.split('_', 1)
            if sheet_key and bs:
                sheet_name = get_sheetname_from_sheetkey(sheet_key)
                need_col_letter = ''
                need_row = -1

                if sheet_name:
                    check_rec_formula_sheet_name = loc_storage_sheet.get(sheet_name)
                    if check_rec_formula_sheet_name:
                        check_rec_unique_key_column = check_rec_formula_sheet_name.get('unique_key_column')
                        if check_rec_unique_key_column:
                            unique_key_column = f'{idx // len_columns}:{col.get("data_type_col")}'
                            check_rec_this_key = check_rec_unique_key_column.get(unique_key_column, None)
                            if check_rec_this_key:
                                need_col_letter = check_rec_this_key.get('Letter')
                        check_rec_set_row_key = check_rec_formula_sheet_name.get('SET_ROW_KEY')
                        if check_rec_set_row_key:
                            check_rec_key_bs = check_rec_set_row_key.get(EnumColumnSettings.KEY_BS)
                            if check_rec_key_bs:
                                check_rec_finded_row = check_rec_key_bs.get(bs)
                                if check_rec_finded_row:
                                    need_row = check_rec_finded_row
                                    if OnlyRow:
                                        return need_row
                if isinstance(need_col_letter, str) and isinstance(need_row, int):
                    if need_col_letter and need_row != -1:
                        return f"'{sheet_name}'!{need_col_letter}{need_row}"
                return None

        # 10030001 -> AB4 или 10030001 -> 4
        def get_cell_addr_by_bs(p, OnlyRow=False):
            # check_rec_sheet_name.get('SET_ROW_KEY').get('key_bs').get('100331977')
            check_rec_set_row_key = check_rec_sheet_name.get('SET_ROW_KEY')
            if check_rec_set_row_key:
                check_rec_key_bs = check_rec_set_row_key.get(EnumColumnSettings.KEY_BS)
                if check_rec_key_bs:
                    need_row = check_rec_key_bs.get(p)
                    if need_row:
                        return need_row if OnlyRow else f'{col_letter}{need_row}'
            return None

        # трансформировать формулу в которой бюджетные статьи имеют ссылки на соответствующие листы, пример:
        # 1_100333000+2_100333001+3_100333002
        # где "1_" - означает ключ листа, а остальное это номер бюджетной статьи
        # Если в формуле не встречается "1_", то текущее значение отличное от нуля считается статьёй
        # с ссылкой на текущий лист из которого и была передана формула для трансформации
        def get_transformed_formula_linked_old(src_formula):
            formula = src_formula
            # исключаем ошибочные формулы в виде ссылок на другие листы (Excel-формулы)
            if '!' in formula:
                src_formula = '0'
            else:
                # сложная регулярка для поиска наших шаблонов вида:
                # 1) 100330001              или 1_100330001
                # 2) (100330001:100330003)  или (1_100330001:1_100330001)
                # 3) {100330001:-1}         или {1_100330001:-1}
                # 4) F6 (Обычная ссылка на ячейку в формулах)

                pattern = r'(\{-?[\d_]+:-?[\d_]+\}|[А-Яа-яA-Za-z_]+\(|\$?[A-Za-z0-9_]+|[^0-9_${}])'
                parts = re.findall(pattern, formula)
                parts = [p for p in parts if p]

                # разделённую на части формулу обработаем, чтобы подменить ключи статей на ссылки ячеек для формул
                src_formula = ''
                for p in parts:
                    FindState = 0
                    if p.startswith('='):
                        # пропускаем обработку такого символа, игнорируя его добавление
                        continue
                    elif p.startswith('$'):
                        p = p.replace('$', '')
                    elif '_' in p:
                        FindState = 1
                        p = get_cell_addr_by_linked_bs(p)
                        if p != None:
                            FindState = 2
                    elif p.isdigit() and int(p) != 0:
                        FindState = 1
                        p = get_cell_addr_by_bs(p)
                        if p != None:
                            FindState = 2
                    elif re.match(r'^\{(-?[\d_]+):(-?[\d_]+)\}$', p):
                        FindState = 1
                        match = re.match(r'^\{(-?[\d_]+):(-?[\d_]+)\}$', p)

                        if match:
                            col = None

                            row_numb = match.group(1)
                            col_offset = match.group(2)
                            try:
                                col_offset = int(col_offset)
                            except ValueError:
                                p = 0
                                pass

                            if '_' in row_numb:
                                # Нет логики обработки ссылочных статей(ссылка ячейки на другой лист)
                                p = 0
                                # loc_sheet_name = ''
                                # row_numb = get_cell_addr_by_linked_bs(row_numb, True)
                                # sheet_key, bs = p.split('_', 1)
                                # if sheet_key and bs:
                                #     loc_sheet_name = get_sheetname_from_sheetkey(sheet_key)
                                #
                                # if loc_sheet_name:
                                #     loc_storage_sheet[loc_sheet_name]['columns_layout'][ + col_offset]
                                #     FindState = 2
                            else:
                                row_numb = get_cell_addr_by_bs(row_numb, True)
                                if row_numb == None:
                                    p = 0
                                else:
                                    col = sheet_columns_layout[idx + col_offset]
                                    p = f'{col.get("Letter")}{row_numb}'
                                    FindState = 2
                    # elif any(re.match(r'^\$?[A-Za-z]{1,2}\$?\d+$', p) for p in parts):
                    elif re.match(r'^\$?[A-Za-z]{1,3}\$?\d+$', p):
                        # удаляем значение столбца далее заменив его на текущий столбец
                        loc_row = re.sub(r'[A-Za-z$]', '', p)
                        p = f'{col_letter}{loc_row}'
                        FindState = 2
                    if FindState == 1:
                        p = 0
                    src_formula = f'{src_formula}{p}'
                pattern = re.compile(r'\b0:0\b|\b0:|:0\b')
                src_formula = pattern.sub('0', src_formula)
                return src_formula
            return '0'

        def get_transformed_formula_linked(src_formula):
            formula = src_formula
            src_formula = ''
            # исключаем ошибочные формулы в виде ссылок на другие листы (Excel-формулы)
            if '!' in formula:
                return '0'
            pattern = re.compile(get_hard_mirror_pattern())

            for match in pattern.finditer(formula):
                token_type = match.lastgroup
                value = match.group()
                new_value = '0'

                # {A1:2}, {1000330001:0} - обработка шаблона
                if token_type == 'CELL_OFFSET':
                    inner_content = value[1:-1]
                    row_numb, col_offset = inner_content.split(':')
                    loc_col = None

                    try:
                        col_offset = int(col_offset)
                        if re.match(r'^\$?[A-Za-z]{1,3}\$?\d+$', row_numb):
                            loc_row = re.sub(r'[A-Za-z$]', '', row_numb)
                            loc_col = sheet_columns_layout[idx + col_offset]
                            new_value = f'{loc_col.get("Letter")}{loc_row}'
                        elif '_' not in row_numb:
                            row_numb = get_cell_addr_by_bs(row_numb, True)
                            if row_numb is not None:
                                loc_col = sheet_columns_layout[idx + col_offset]
                                new_value = f'{loc_col.get("Letter")}{row_numb}'
                        else:
                            # логики для обработки ссылочной статьи нет
                            pass

                    except Exception as e:
                        pass

                # Округл - Имя функции формулы-xls
                # A1 - ссылка на ячейку
                # 1000330001 - ключ статьи
                # 1_1000330001 - ключ статьи со ссылкой на лист
                elif token_type == 'FUNC_OR_VAR':
                    end_index = match.end()
                    next_char = formula[end_index] if end_index < len(formula) else ''
                    # Проверяем, если попало имя функции, то просто передаём его дальше без обработки
                    if next_char == '(':
                        new_value = f'{value}'
                    elif re.match(r'^\$?[A-Za-z]{1,3}\$?\d+$', value):
                        # удаляем значение столбца далее заменив его на текущий столбец
                        loc_row = re.sub(r'[A-Za-z$]', '', value)
                        new_value = f'{col_letter}{loc_row}'
                    else:
                        if '_' in value:
                            new_value = get_cell_addr_by_linked_bs(value)
                            if new_value is None:
                                new_value = '0'
                        else:
                            new_value = get_cell_addr_by_bs(value)
                            if new_value is None:
                                new_value = value

                # экранирование для $
                elif token_type == 'ESCAPED':
                    new_value = value.replace('$', '')

                # обычный элемент
                else:
                    new_value = value

                src_formula = f'{src_formula}{new_value}'
            return src_formula if src_formula != '' else '0'

        finded_map = None
        wb = cell.parent.parent

        check_rec_sheet_name = loc_storage_sheet.get(cell.parent.title)
        if check_rec_sheet_name:
            check_rec_set_row_idx = check_rec_sheet_name.get('SET_ROW_IDX')
            if check_rec_set_row_idx:
                check_rec_index = check_rec_set_row_idx.get(cell.row)
                if check_rec_index:
                    check_rec_some_key = check_rec_index.get(ColumnSettings)
                    if check_rec_some_key:
                        check_rec_value = check_rec_some_key.get('value')
                        if check_rec_value:
                            finded_map = check_rec_value
        if finded_map:
            src_formula = get_transformed_formula_linked(finded_map)
            # openpyxl не умеет в русские формулы поэтому конвертируем из русской формулы в английскую
            # чтобы на выходе в Exel или Р7-офис формула автоматически пересчитывались (работали)
            set_value_cell(cell, formula_translator.convert_russian_formula(src_formula), ColumnCellType)
            return
            # if finded_map.get('bs') == cell_bs.value and finded_map.get('calc') != None and finded_map.get('calc') != '':
            #     src_formula = get_transformed_formula_linked(finded_map.get('calc'))
            #     # openpyxl не умеет в русские формулы поэтому конвертируем из русской формулы в английскую
            #     # чтобы на выходе в Exel или Р7-офис формула автоматически пересчитывались (работали)
            #     set_value_cell(cell, ft.convert_russian_formula(src_formula), ColumnType)
            #     return
            # elif finded_map.get('formula', ''):
            #     src_formula = get_transformed_formula_linked(finded_map.get('formula'))
            #
            #     # openpyxl не умеет в русские формулы поэтому конвертируем из русской формулы в английскую
            #     # чтобы на выходе в Exel или Р7-офис формула автоматически пересчитывались (работали)
            #     set_value_cell(cell, ft.convert_russian_formula(src_formula), ColumnType)
            #     return
        if query_res:
            if query_res[0] == 0:
                set_value_cell(cell, 0)
                return
            else:
                for row in query_res:
                    try:
                        if int(row.bs) == int(query_cell.value) and col["calmonth"] == row.calmonth:
                            set_value_cell(cell, row.sum, ColumnCellType)
                            return
                    except Exception as e:
                        pass
        set_value_cell(cell, None, ColumnCellType)
        return

    def get_filled_formula_by_col(col, l_col: list):
        if 'FormulaLink' in col and 'total' in col.get('FormulaLink'):
            FormulaLink = col.get('FormulaLink')
            total = FormulaLink.get('total')
            if total == len(l_col):
                # src_formula = f"=ОКРУГЛ({letter_col}{i}/{letter_col}{FirstRowData}*100;6)"
                src_formula = FormulaLink.get('Formula')  # [1:]
                cell_address = []
                for i in range(1, total + 1):
                    Letter = FormulaLink.get(f'col{i}').get('Letter')
                    row = l_col[i - 1]
                    cell_address.append(f'{Letter}{row}')
                src_formula = src_formula.format(*cell_address)
                return src_formula
        return ''

    cell_bs = None
    cell_calc = None

    SetRowIdx = loc_storage_sheet[sheet.title]['SET_ROW_IDX']
    key_bs = loc_storage_sheet[sheet.title]['columns_settings'][EnumColumnSettings.KEY_BS]
    key_spec = loc_storage_sheet[sheet.title]['columns_settings'][EnumColumnSettings.KEY_SPEC]
    key_input = loc_storage_sheet[sheet.title]['columns_settings'][EnumColumnSettings.KEY_INPUT]
    formula_month = loc_storage_sheet[sheet.title]['columns_settings'][EnumColumnSettings.FORMULA_MONTH]
    formula_poo = loc_storage_sheet[sheet.title]['columns_settings'][EnumColumnSettings.FORMULA_POO]

    prepared_column_offset_1 = -1
    prepared_column_offset_2 = -1
    if type_sheet == 'type_factory':
        prepared_column_offset_1 = key_bs
        prepared_column_offset_2 = formula_month
    elif type_sheet == 'type_summary_rep':
        prepared_column_offset_1 = formula_month  # key_bs
        prepared_column_offset_2 = formula_month

    FirstRowWithBS = -1
    check_rec_title = loc_storage_sheet.get(sheet.title)
    if check_rec_title:
        check_rec_row_with_bs = check_rec_title.get('FirstRowWithBS')
        if check_rec_row_with_bs:
            FirstRowWithBS = check_rec_row_with_bs

    if FirstRowWithBS == -1:
        FirstRowWithBS = FirstRowData

    # for i in range(FirstRowData, LastRowData + 1):
    for i, v in SetRowIdx.items():
        cell_key_spec = sheet.cell(row=i, column=key_spec)
        cell_key_input = sheet.cell(row=i, column=key_input)
        cell_bs = sheet.cell(row=i, column=prepared_column_offset_1)
        cell_calc = sheet.cell(row=i, column=prepared_column_offset_2)



        if not (isinstance(cell_bs.value, str) or isinstance(cell_bs.value, int) or \
                cell_key_spec.value is not None or cell_key_input.value is not None):
            continue

        if cell_key_input.value is not None:
            for idx, col in enumerate(sheet_columns_layout):
                col_letter = col.get('Letter')
                data_type_col = col['data_type_col']
                SrcKey = col['SrcKey']
                ColumnType = col['ColumnType']

                if 'Byear' in col['type']:
                    if ColumnType == 'Byear':
                        SrcKeyByear = col['SrcKeyByear']
                        cell_val = sheet[f'{col_letter}{i}']
                        set_calc_cell_val(cell_val, cell_key_input, col_letter,
                                          query_res_for_column[SrcKey], col, idx,
                                          ColumnCellType=EnumCellType.POSITIVE_NEGATIVE)
                elif 'year' in col['type']:
                    src_formula = ''
                    if ColumnType == 'Selected':
                        q1 = get_col_letter_by_type('Q1', data_type_col)
                        q2 = get_col_letter_by_type('Q2', data_type_col)
                        q3 = get_col_letter_by_type('Q3', data_type_col)
                        q4 = get_col_letter_by_type('Q4', data_type_col)
                        set_value_cell(sheet[f'{col_letter}{i}'], f'={q1}{i}+{q2}{i}+{q3}{i}+{q4}{i}')
                    elif ColumnType == 'PercentOfOutput':
                        cell_val = sheet[f'{col_letter}{i}']
                        set_calc_cell_val(cell_val, cell_key_input, col_letter,
                                          query_res_for_column[SrcKey], col, idx,
                                          ColumnSettings=EnumColumnSettings.FORMULA_POO)

                        # src_formula = get_filled_formula_by_col(col,[i,FirstRowWithBS])
                        #
                        # set_value_cell(sheet[f"{col_letter}{i}"],
                        #                formula_translator.convert_russian_formula(src_formula))
                    elif ColumnType in ('SecondMinusFirst', 'PercentOfComplete'):
                        src_formula = get_filled_formula_by_col(col, [i, i])
                        if ColumnType == 'SecondMinusFirst':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.POSITIVE_NEGATIVE)
                        elif ColumnType == 'PercentOfComplete':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.PERCENT)
                elif "Q" in col["type"]:
                    q_num = col["type"][1]

                    if ColumnType == 'Selected':
                        m1 = get_col_letter_by_type(f"M{q_num}_1", data_type_col)
                        m2 = get_col_letter_by_type(f"M{q_num}_2", data_type_col)
                        m3 = get_col_letter_by_type(f"M{q_num}_3", data_type_col)
                        set_value_cell(sheet[f"{col_letter}{i}"], f"={m1}{i}+{m2}{i}+{m3}{i}")
                    elif ColumnType == 'PercentOfOutput':
                        cell_val = sheet[f"{col_letter}{i}"]
                        set_calc_cell_val(cell_val, cell_key_input, col_letter,
                                          query_res_for_column[SrcKey], col, idx,
                                          ColumnSettings=EnumColumnSettings.FORMULA_POO)
                    elif ColumnType in ('SecondMinusFirst', 'PercentOfComplete'):
                        src_formula = get_filled_formula_by_col(col, [i, i])
                        if ColumnType == 'SecondMinusFirst':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.POSITIVE_NEGATIVE)
                        elif ColumnType == 'PercentOfComplete':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.PERCENT)
                elif "M" in col["type"]:
                    if ColumnType == 'Selected':
                        cell_val = sheet[f"{col_letter}{i}"]
                        set_calc_cell_val(cell_val, cell_key_input, col_letter,
                                          query_res_for_column[SrcKey], col, idx,
                                          ColumnCellType = EnumCellType.CELL_INPUT)
                    elif ColumnType == 'PercentOfOutput':
                        cell_val = sheet[f"{col_letter}{i}"]
                        set_calc_cell_val(cell_val, cell_key_input, col_letter,
                                          query_res_for_column[SrcKey], col, idx,
                                          ColumnSettings=EnumColumnSettings.FORMULA_POO)
                    elif ColumnType in ('SecondMinusFirst', 'PercentOfComplete'):
                        src_formula = get_filled_formula_by_col(col, [i, i])
                        if ColumnType == 'SecondMinusFirst':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.POSITIVE_NEGATIVE)
                        elif ColumnType == 'PercentOfComplete':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.PERCENT)
        # расчёт для специальных строк
        elif "T" in str(cell_key_spec.value):
            if cell_calc and cell_calc.value is not None:
                for idx, col in enumerate(sheet_columns_layout):
                    col_letter = col.get('Letter')
                    data_type_col = col["data_type_col"]
                    SrcKey = col["SrcKey"]
                    ColumnType = col["ColumnType"]
                    cell_val = sheet[f"{col_letter}{i}"]

                    if ColumnType == 'Selected':
                        set_calc_cell_val(cell_val, cell_bs, col_letter,
                                          query_res_for_column[SrcKey], col, idx)
                    elif ColumnType == 'PercentOfOutput':
                        cell_val = sheet[f"{col_letter}{i}"]
                        set_calc_cell_val(cell_val, cell_bs, col_letter,
                                          query_res_for_column[SrcKey], col, idx)
                        # ColumnSettings=EnumColumnSettings.FORMULA_POO)
                        # src_formula = get_filled_formula_by_col(col, [i, FirstRowWithBS])
                        # set_value_cell(sheet[f"{col_letter}{i}"],
                        #                formula_translator.convert_russian_formula(src_formula))
                    elif ColumnType in ('SecondMinusFirst', 'PercentOfComplete'):
                        src_formula = get_filled_formula_by_col(col, [i, i])
                        if ColumnType == 'SecondMinusFirst':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.POSITIVE_NEGATIVE)
                        elif ColumnType == 'PercentOfComplete':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.PERCENT)
        # Проверочные строки
        elif "V" in str(cell_key_spec.value):
            for idx, col in enumerate(sheet_columns_layout):
                col_letter = col.get('Letter')
                data_type_col = col["data_type_col"]
                SrcKey = col["SrcKey"]
                ColumnType = col["ColumnType"]
                if ColumnType == 'Selected':
                    cell_val = sheet[f"{col_letter}{i}"]
                    set_calc_cell_val(cell_val, cell_bs, col_letter,
                                      query_res_for_column[SrcKey], col, idx, ColumnCellType=EnumCellType.FONT_CHECKER1)
        # Проверочные строки #2
        elif "S1" in str(cell_key_spec.value):
            for idx, col in enumerate(sheet_columns_layout):
                col_letter = col.get('Letter')
                data_type_col = col["data_type_col"]
                SrcKey = col["SrcKey"]
                ColumnType = col["ColumnType"]
                if ColumnType == 'Selected' and SrcKey == 0 and "M" in col["type"]:
                    cell_val = sheet[f"{col_letter}{i}"]
                    set_calc_cell_val(cell_val, cell_bs, col_letter,
                                      query_res_for_column[SrcKey], col, idx,
                                      ColumnCellType=EnumCellType.FONT_CHECKER1)
        # Подсчёт строк для столбцов отклонение #2
        elif "S2" in str(cell_key_spec.value):
            for idx, col in enumerate(sheet_columns_layout):
                col_letter = col.get('Letter')
                data_type_col = col["data_type_col"]
                SrcKey = col["SrcKey"]
                ColumnType = col["ColumnType"]
                if ColumnType == 'SecondMinusFirst':
                    cell_val = sheet[f"{col_letter}{i}"]
                    set_calc_cell_val(cell_val, cell_bs, col_letter,
                                      query_res_for_column[SrcKey], col, idx,
                                      ColumnCellType=EnumCellType.FONT_CHECKER1)
        # Не справочные строки
        # забираем только не пустые
        elif cell_bs and cell_bs.value is not None and cell_bs.value != 0:
            # (str(cell_bs.value).isdigit() or '_' in str(cell_bs.value) or \
            #  (str(cell_bs.value).startswith('='))): # логика для формул на листах СВОД
            for idx, col in enumerate(sheet_columns_layout):
                col_letter = col.get('Letter')
                data_type_col = col['data_type_col']
                SrcKey = col['SrcKey']
                ColumnType = col['ColumnType']

                if 'Byear' in col['type']:
                    if ColumnType == 'Byear':
                        SrcKeyByear = col['SrcKeyByear']
                        cell_val = sheet[f'{col_letter}{i}']
                        set_calc_cell_val(cell_val, cell_bs, col_letter,
                                          query_res_for_column[SrcKey], col, idx,
                                          ColumnCellType=EnumCellType.POSITIVE_NEGATIVE)
                elif 'year' in col['type']:
                    src_formula = ''
                    if ColumnType == 'Selected':
                        q1 = get_col_letter_by_type('Q1', data_type_col)
                        q2 = get_col_letter_by_type('Q2', data_type_col)
                        q3 = get_col_letter_by_type('Q3', data_type_col)
                        q4 = get_col_letter_by_type('Q4', data_type_col)
                        set_value_cell(sheet[f'{col_letter}{i}'], f'={q1}{i}+{q2}{i}+{q3}{i}+{q4}{i}')
                    elif ColumnType == 'PercentOfOutput':
                        cell_val = sheet[f'{col_letter}{i}']
                        set_calc_cell_val(cell_val, cell_bs, col_letter,
                                          query_res_for_column[SrcKey], col, idx,
                                          ColumnSettings=EnumColumnSettings.FORMULA_POO)

                        # src_formula = get_filled_formula_by_col(col,[i,FirstRowWithBS])
                        #
                        # set_value_cell(sheet[f"{col_letter}{i}"],
                        #                formula_translator.convert_russian_formula(src_formula))
                    elif ColumnType in ('SecondMinusFirst', 'PercentOfComplete'):
                        src_formula = get_filled_formula_by_col(col, [i, i])
                        if ColumnType == 'SecondMinusFirst':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.POSITIVE_NEGATIVE)
                        elif ColumnType == 'PercentOfComplete':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.PERCENT)
                elif "Q" in col["type"]:
                    q_num = col["type"][1]

                    if ColumnType == 'Selected':
                        m1 = get_col_letter_by_type(f"M{q_num}_1", data_type_col)
                        m2 = get_col_letter_by_type(f"M{q_num}_2", data_type_col)
                        m3 = get_col_letter_by_type(f"M{q_num}_3", data_type_col)
                        set_value_cell(sheet[f"{col_letter}{i}"], f"={m1}{i}+{m2}{i}+{m3}{i}")
                    elif ColumnType == 'PercentOfOutput':
                        cell_val = sheet[f"{col_letter}{i}"]
                        set_calc_cell_val(cell_val, cell_bs, col_letter,
                                          query_res_for_column[SrcKey], col, idx,
                                          ColumnSettings=EnumColumnSettings.FORMULA_POO)
                        # src_formula = get_filled_formula_by_col(col, [i, FirstRowWithBS])
                        # set_value_cell(sheet[f"{col_letter}{i}"],
                        #                formula_translator.convert_russian_formula(src_formula),)
                    elif ColumnType in ('SecondMinusFirst', 'PercentOfComplete'):
                        src_formula = get_filled_formula_by_col(col, [i, i])
                        if ColumnType == 'SecondMinusFirst':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.POSITIVE_NEGATIVE)
                        elif ColumnType == 'PercentOfComplete':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.PERCENT)
                elif "M" in col["type"]:
                    if ColumnType == 'Selected':
                        cell_val = sheet[f"{col_letter}{i}"]
                        set_calc_cell_val(cell_val, cell_bs, col_letter,
                                          query_res_for_column[SrcKey], col, idx)
                    elif ColumnType == 'PercentOfOutput':
                        cell_val = sheet[f"{col_letter}{i}"]
                        set_calc_cell_val(cell_val, cell_bs, col_letter,
                                          query_res_for_column[SrcKey], col, idx,
                                          ColumnSettings=EnumColumnSettings.FORMULA_POO)
                        # src_formula = get_filled_formula_by_col(col, [i, FirstRowWithBS])
                        # set_value_cell(sheet[f"{col_letter}{i}"],
                        #                formula_translator.convert_russian_formula(src_formula))
                    elif ColumnType in ('SecondMinusFirst', 'PercentOfComplete'):
                        src_formula = get_filled_formula_by_col(col, [i, i])
                        if ColumnType == 'SecondMinusFirst':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.POSITIVE_NEGATIVE)
                        elif ColumnType == 'PercentOfComplete':
                            set_value_cell(sheet[f"{col_letter}{i}"],
                                           formula_translator.convert_russian_formula(src_formula),
                                           EnumCellType.PERCENT)


# ============================================================================================
def get_report_template(id):
    template_name = g_report_template_name
    path_template = str(Path(uf.main_folder) / Path(uf.file_folder) / Path(template_name))

    # Получаем рабочую книгу из шаблона
    wb = openpyxl.load_workbook(path_template)

    SHEETS_SETTINGS = get_sheets_settings()

    if not SHEETS_SETTINGS:
        return uf.get_msg_struct(uf.EnumMsg.SETTINGS_FOR_REPORT_NOT_FOUND)

    sheet_id_all = [item.get('sheet_id') for item in SHEETS_SETTINGS]
    for loc_sheet_id in sheet_id_all:
        try:
            sheet = get_sheet_from_sheet_id(wb,loc_sheet_id)
            if sheet:
                if id != str(loc_sheet_id):
                    sheet.sheet_state = 'veryHidden'
                else:
                    sheet.sheet_state = 'visible'
        except Exception as e:
            continue

    # Создаём буфер для наполнения
    buffer = io.BytesIO()

    wb.save(buffer)
    # Откатываем курсор в самое начало
    buffer.seek(0)
    # имя файла
    filename = template_name

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============================================================================================
def upload_report_template(sheet_id: str, file_storage: FileStorage):
    # Проверка, замена левых символов в имени файла
    # safe_filename = secure_filename(file_storage.filename)

    try:
        file_bytes = file_storage.stream.read()
        file_stream = io.BytesIO(file_bytes)
        wb = load_workbook(filename=file_stream, read_only=False)

    except Exception as e:
        return uf.get_msg_struct(uf.EnumMsg.ERROR_OPEN_TEMPLATE, str(e))

    SHEETS_SETTINGS = get_sheets_settings()
    if not SHEETS_SETTINGS:
        return uf.get_msg_struct(uf.EnumMsg.SETTINGS_FOR_REPORT_NOT_FOUND)

    sheets_settings_report_static = get_sheet_list_by_field(SHEETS_SETTINGS,'type_sheet', 'STATIC')
    sheets_settings_upload = get_sheet_list_by_field(SHEETS_SETTINGS,'upload', '1')

    try:
        # ==============================================================================================================
        def prepare_sheets(sheet_list, name_rng):
            def_range_bs_list = []
            def_range_ik_list = []
            for sheet_id in sheet_list:
                try:
                    def_range_bs_list.append(wb.defined_names[f'{name_rng[0]}{sheet_id}'])
                    def_range_ik_list.append(wb.defined_names[f'{name_rng[1]}{sheet_id}'])
                except Exception as e:
                    return uf.get_msg_struct(uf.EnumMsg.ERROR_VALID_NEW_TEMPLATE)
            # ===============================================================================
            for idx, def_range_bs in enumerate(def_range_bs_list):
                def_range_ik = def_range_ik_list[idx]
                sheet = None
                set_row_right_col_index = None
                start_row = None
                start_col = None
                end_col = None
                amount_col = None

                for sheet_name_rng_ik, cell_coordinates_rng_ik in def_range_ik.destinations:
                    _, start_row, _, _ = range_boundaries(cell_coordinates_rng_ik)

                for sheet_name_rng_bs, cell_coordinates_rng_bs in def_range_bs.destinations:
                    _, _, set_row_right_col_index, _ = range_boundaries(cell_coordinates_rng_bs)
                    sheet = wb[sheet_name_rng_bs]

                if sheet is not None and set_row_right_col_index is not None and start_row is not None:
                    start_col = set_row_right_col_index + 2
                    end_col = sheet.max_column + 1
                    amount_col = end_col - start_col

                    start_col_letter = get_column_letter(start_col)
                    finded_start_col_letter = False

                    # Очистка сгенерированных столбцов:
                    #   из-за особенностей работы библы openpyxl некоторые очистки приходится делать дополнительно
                    #   удаление группировки, объединённые ячейки, ширина столбцов
                    #   если правильно обнаружил, то 8.43 ширина является значением по умолчанию

                    # Из-за того, что в column_dimensions столбцы могут храниться в хаотичном порядке
                    # небходимо их сначала отсортировать, чтобы задействовать только сгенерированные столбцы
                    sorted_dims = sorted(
                        sheet.column_dimensions.values(), key=lambda x: column_index_from_string(x.index)
                        if hasattr(x, 'index') and x.index else
                        column_index_from_string(x.key)
                    )

                    for col_dim in sorted_dims:
                        col_idx = column_index_from_string(
                            col_dim.index if hasattr(col_dim, 'index') and col_dim.index else col_dim.key)

                        if col_idx >= start_col:
                            col_dim.width = 8.43
                            col_dim.outline_level = 0
                            col_dim.hidden = False

                    # Очистка столбцов
                    sheet.delete_cols(idx=start_col, amount=amount_col)

                    # Из-за особенностей openpyxl приходится дополнительно делать очистку по merged ячейкам(диапазонам)
                    merged_range = list(sheet.merged_cells.ranges)
                    for m_range in merged_range:
                        merged_start_col, _, merged_end_col, _ = m_range.bounds
                        if merged_start_col <= end_col and merged_end_col >= start_col:
                            try:
                                sheet.merged_cells.remove(m_range)
                            except ValueError:
                                pass
        # ==============================================================================================================

        # Старый способ проверки загружаемого шаблона
        # factories_all = [row.id for row in uf.get_data_from_query("SELECT id FROM tab_factories_d816_4")]
        # reports_all_list = [value.get('index') for key, value in generating_report_settings.items()]
        #
        # if factories_all and reports_all_list:
        #     prepare_sheets(factories_all, ['_SET_ROW', '_INTERNAL_KEY'])
        #     prepare_sheets(reports_all_list, ['_SUM_REP_SET_ROW', '_SUM_REP_INTERNAL_KEY'])
        # else:
        #     return uf.get_msg_struct(uf.EnumMsg.SETTINGS_FOR_REPORT_NOT_FOUND)

        def check_this_sheet_id_template(loc_sheet_id):
            # ==========================================================================================================
            # SHEET_ID
            sheet_id_data = get_data_from_named_range_name(wb, f'_SHEET_ID{loc_sheet_id}')
            if sheet_id_data.get('Exec', False) == False:
                return 1
            # ==========================================================================================================
            nr_sheet_id_sheet_name = sheet_id_data.get('sheet_name')
            # Именованная диапазоны, которые должны быть обязаны на каждом вычислительном листе
            # _set_row = get_named_rng_partial_name(wb, '_SET_ROW')
            # _internal_key = get_named_rng_partial_name(wb, '_INTERNAL_KEY')
            nr_set_row = get_named_range_from_sheet_id_and_nr_name(wb, loc_sheet_id, '_SET_ROW')
            nr_internal_key = get_named_range_from_sheet_id_and_nr_name(wb, loc_sheet_id, '_INTERNAL_KEY')


            if nr_set_row is None or nr_internal_key is None:
                return 2
            _set_row_data = get_data_from_named_range(wb, nr_set_row)
            _internal_key_data = get_data_from_named_range(wb, nr_internal_key)

            sheet_name_set_row = nr_sheet_id_sheet_name
            set_row_left_col_index = _set_row_data.get('min_col')
            set_row_right_col_index = _set_row_data.get('max_col')
            sheet = sheet_id_data.get('sheet')

            # sheet.sheet_state = 'veryHidden'

            FirstRowData = _internal_key_data.get('min_row')
            first_row = FirstRowData - 3  # всего строк для заголовков 3
            FirstRowData += 1
            last_row = sheet.max_row


            # ==========================================================================================================
            # 22 - ключ для "Отчет по КПД ДО" у него статичные столбцы
            if loc_sheet_id not in sheets_settings_report_static:
                start_col = set_row_right_col_index + 2
                end_col = sheet.max_column + 1
                amount_col = end_col - start_col

                start_col_letter = get_column_letter(start_col)
                finded_start_col_letter = False

                # Очистка сгенерированных столбцов:
                #   из-за особенностей работы библы openpyxl некоторые очистки приходится делать дополнительно
                #   удаление группировки, объединённые ячейки, ширина столбцов
                #   если правильно обнаружил, то 8.43 ширина является значением по умолчанию

                # Из-за того, что в column_dimensions столбцы могут храниться в хаотичном порядке
                # небходимо их сначала отсортировать, чтобы задействовать только сгенерированные столбцы
                sorted_dims = sorted(
                    sheet.column_dimensions.values(), key=lambda x: column_index_from_string(x.index)
                    if hasattr(x, 'index') and x.index else
                    column_index_from_string(x.key)
                )

                for col_dim in sorted_dims:
                    col_idx = column_index_from_string(
                        col_dim.index if hasattr(col_dim, 'index') and col_dim.index else col_dim.key)

                    if col_idx >= start_col:
                        col_dim.width = 8.43
                        col_dim.outline_level = 0
                        col_dim.hidden = False

                # Очистка столбцов
                sheet.delete_cols(idx=start_col, amount=amount_col)

                # Из-за особенностей openpyxl приходится дополнительно делать очистку по merged ячейкам(диапазонам)
                merged_range = list(sheet.merged_cells.ranges)
                for m_range in merged_range:
                    merged_start_col, _, merged_end_col, _ = m_range.bounds
                    if merged_start_col <= end_col and merged_end_col >= start_col:
                        try:
                            sheet.merged_cells.remove(m_range)
                        except ValueError:
                            pass
                # ======================================================================================================
                # Процесс генерации ключей для ввода данных
                # 21 - Генерация пока что только для отчёта Баланс ЗС
                if loc_sheet_id not in sheets_settings_upload:
                    columns_settings = get_common_column_settings(set_row_left_col_index)
                    dict_exists_key_input = []
                    dict_new_key_input = []
                    dict_input_data_tab = []
                    for i in range(FirstRowData, last_row+1):
                        cell_key_input = sheet.cell(row=i, column=columns_settings[EnumColumnSettings.KEY_INPUT])
                        cell_key_bs = sheet.cell(row=i, column=columns_settings[EnumColumnSettings.KEY_BS])
                        cell_formula_month = sheet.cell(row=i, column=columns_settings[EnumColumnSettings.FORMULA_MONTH])
                        if cell_key_input.value is not None:
                            dict_exists_key_input.append(cell_key_input.value)
                        elif cell_key_input.value is None and cell_key_bs.value is None and cell_formula_month.value is None:
                            # Если ключ для ввода пустой и нет ключа для статьи и нет ключа для месяца, то
                            dict_new_key_input.append(
                                {
                                    'cell' : cell_key_input,
                                    'key_bs' : None,
                                }
                            )

                    # генерируем ключи для ввода данных, если есть пустые ячейки для ключа "ввод данных"
                    if dict_new_key_input:
                        # next_value = max(dict_exists_key_input, default=0)
                        next_value = uf.get_dict_data_from_query(f"""
                            SELECT
                                coalesce(max(bs),0) as max
                            FROM
                                tab_integ_get_preu_mirror_d816_4
                            WHERE
                                sheet_id = {loc_sheet_id}
                        """)[0].get('max',0)
                        for item in dict_new_key_input:
                            next_value += 1
                            item['key_bs'] = next_value
                            dict_input_data_tab.append(
                                {
                                    'sheet_id' : loc_sheet_id,
                                    'key_bs' : item['key_bs'],
                                }
                            )
                    # Если есть сгенерированные ключи, то попробуем их записать в БД
                    if dict_input_data_tab:
                        dict_main_input_data_tab[loc_sheet_id] = {
                            'dict_input_data_tab' : dict_input_data_tab,
                            'dict_new_key_input' : dict_new_key_input
                        }
                # ======================================================================================================
            else:
                pass
            return 0

        # sheet_id_all = [{'id' : row.sheet_id, 'name' : row.name_sheet }for row in uf.get_data_from_query("""
        #     SELECT
        #         sheet_id,
        #         name_sheet
        #     FROM
        #         tab_sheet_id_list_d816_4
        #     ORDER BY
        #         SHEET_ID
        # """)]
        sheet_id_all = [
            {
                'id' : item.get('sheet_id'),
                'name' : item.get('name_sheet')
            }
            for item in SHEETS_SETTINGS
        ]
        # sheet_id_all = [
        #     { 'id' : 1, 'name' : 'sheet 1' },
        #     { 'id' : 2, 'name' : 'sheet 2' },
        #     { 'id' : 20, 'name' : 'sheet 20' },
        #     { 'id' : 21, 'name' : 'sheet 21' },
        #     { 'id' : 22, 'name' : 'sheet 22' },
        # ]
        sheet_id_all = [
            {'id': 21, 'name': 'Баланс ЗС'}, # 21 -Генерация пока что только для Баланс ЗС
        ]

        dict_main_input_data_tab = {}
        for row in sheet_id_all:
            id = row.get('id')
            name = row.get('name')
            check_result = check_this_sheet_id_template(id)
            if check_result != 0:
                return uf.get_msg_struct(uf.EnumMsg.ERROR_VALID_NEW_TEMPLATE, f"№e{check_result} id {id} - {name}")

        for item in dict_main_input_data_tab.values():
            save_input_data_key = insert_sheet_key_in_tab(item.get('dict_input_data_tab', None))
            if save_input_data_key:
                for row in item.get('dict_new_key_input', None):
                    row.get('cell').value = row.get('key_bs', None)
        path_template = str(Path(uf.main_folder) / Path(uf.file_folder))

        base_template_dir = os.environ.get("TEMPLATE_DIR", path_template)

        target_filename = f"{g_report_template_name}".lower()
        target_path = os.path.join(base_template_dir, target_filename)

        dir_to_create = os.path.dirname(target_path)
        if not os.path.exists(dir_to_create):
            try:
                os.makedirs(dir_to_create, mode=0o755, exist_ok=True)
            except PermissionError:
                wb.close()
                return uf.get_msg_struct(uf.EnumMsg.ERROR_PERMISSION_CREATE_DIR_LINUX, dir_to_create)

        # Перезапись существующего файла шаблона
        if os.path.exists(target_path) and not os.access(target_path, os.W_OK):
            wb.close()
            return uf.get_msg_struct(uf.EnumMsg.ERROR_PERMISSION_OVERWRITE_FILE)

        # Сохранение файла
        wb.save(target_path)
        wb.close()

        return uf.get_msg_struct(uf.EnumMsg.SUCCESS)

    except Exception as e:
        if 'wb' in locals():
            wb.close()

        _, _, exc_tb = sys.exc_info()
        tb_info = traceback.extract_tb(exc_tb)[-1]

        error_msg = (
            f"Error: {e} | "
            f"File: {tb_info.filename} | "
            f"Line: {tb_info.lineno} | "
            f"Func: {tb_info.name}"
        )
        return uf.get_msg_struct(uf.EnumMsg.ERROR_SAVE_OR_PROC_TEMPLATE, error_msg)
# ============================================================================================
# ============================================================================================
def upload_report(sheet_id: str, file_storage: FileStorage):
    # Проверка, замена левых символов в имени файла
    # safe_filename = secure_filename(file_storage.filename)

    try:
        file_bytes = file_storage.stream.read()
        file_stream = io.BytesIO(file_bytes)
        wb = load_workbook(filename=file_stream, read_only=False)

    except Exception as e:
        return uf.get_msg_struct(uf.EnumMsg.ERROR_OPEN_TEMPLATE, str(e))

    SHEETS_SETTINGS = get_sheets_settings()
    if not SHEETS_SETTINGS:
        return uf.get_msg_struct(uf.EnumMsg.SETTINGS_FOR_REPORT_NOT_FOUND)

    sheets_settings_report_static = get_sheet_list_by_field(SHEETS_SETTINGS,'type_sheet', 'STATIC')
    sheets_settings_with_upload = get_sheet_list_by_field(SHEETS_SETTINGS,'upload', '1')

    try:
        def save_data_this_sheet_id(loc_sheet_id):
            # ==========================================================================================================
            # SHEET_ID
            nr_sheet_id = get_data_from_named_range_name(wb, f'_SHEET_ID{loc_sheet_id}')
            if nr_sheet_id.get('Exec', False) == False:
                return 1
            # ==========================================================================================================
            nr_sheet_id_sheet_name = nr_sheet_id.get('sheet_name')
            # Именованная диапазоны, которые должны быть обязаны на каждом вычислительном листе
            # _set_row = get_named_rng_partial_name(wb, '_SET_ROW')
            # _internal_key = get_named_rng_partial_name(wb, '_INTERNAL_KEY')

            _set_row = get_named_range_from_sheet_id_and_nr_name(wb, loc_sheet_id, '_SET_ROW')
            _internal_key = get_named_range_from_sheet_id_and_nr_name(wb, loc_sheet_id, '_INTERNAL_KEY')

            if _set_row is None or _internal_key is None:
                return 2
            nr_set_row = get_data_from_named_range(wb, _set_row)
            nr_internal_key = get_data_from_named_range(wb, _internal_key)

            sheet_name_set_row = nr_sheet_id_sheet_name
            set_row_left_col_index = nr_set_row.get('min_col')
            set_row_right_col_index = nr_set_row.get('max_col')
            sheet = nr_sheet_id.get('sheet')

            FirstRowData = nr_internal_key.get('min_row')
            first_row = FirstRowData - 3  # всего строк для заголовков 3
            FirstRowData += 1
            last_row = sheet.max_row

            start_col = nr_set_row.get('max_col') + 2
            end_col = sheet.max_column


            # ==========================================================================================================
            # Исключаем отчёты со статичными столбцами
            # в будущем возможно такая особенность будет нужна и её возможно реализовать, но
            # сейчас такой реализации нет
            if loc_sheet_id not in sheets_settings_report_static:
                # ======================================================================================================
                # Процесс генерации ключей для ввода данных
                columns_settings = get_common_column_settings(set_row_left_col_index)
                list_cell_key_input = []
                new_insert_rows = []
                for i in range(FirstRowData, last_row + 1):
                    cell_key_input = sheet.cell(row=i, column=columns_settings[EnumColumnSettings.KEY_INPUT])
                    if cell_key_input.value is not None:
                        list_cell_key_input.append(cell_key_input)
                if list_cell_key_input:
                    for cell_key_input in list_cell_key_input:
                        loc_row = cell_key_input.row
                        for col in range(start_col, end_col+1):
                            cell_ik = sheet.cell(row=nr_internal_key.get('min_row'), column=col)
                            if cell_ik.value is not None and re.match(EnumInternalKeyPatterns.MONTH, cell_ik.value):
                                InternalKeys = SplitInternalKey(cell_ik.value)
                                if InternalKeys != None:
                                    cell_sum_value = sheet.cell(row=loc_row, column=col).value
                                    if cell_sum_value is not None:
                                        InternalKeys['bs'] = cell_key_input.value # Бюджетная статья
                                        InternalKeys['sheet_id'] = loc_sheet_id
                                        InternalKeys['sum'] = cell_sum_value
                                        new_insert_rows.append(build_dict_row(InternalKeys, EnumFuncModuParameter.proizv_pererab.value))
                if new_insert_rows:
                    if insert_get_preu_mirror_in_tab(new_insert_rows):
                        return 0
                    else:
                        return 3
                # ======================================================================================================
            else:
                pass
            return 0

        # sheet_id_all = [
        #     {'id' : row.id, 'name' : row.name }
        #     for row in uf.get_data_from_query("SELECT sheet_id as id, name_display as name FROM tab_sheet_id_list_d816_4")
        #     if row.id == 21 # 21 - принудительно только для Баланс ЗС
        # ]
        sheet_id_all = sheets_settings_with_upload
        sheet_id_list_all = [item.get('sheet_id') for item in sheet_id_all]
        if int(sheet_id) in sheet_id_list_all:
            row = next((item for item in sheet_id_all if item['sheet_id'] == int(sheet_id)), None)
            result_save = save_data_this_sheet_id(row.get('sheet_id'))
            if result_save == 0:
                return uf.get_msg_struct(uf.EnumMsg.ERROR_VALID_NEW_TEMPLATE, f"№e{result_save} id {row.get('sheet_id')} - {row.get('name')}")

        wb.close()

        return uf.get_msg_struct(uf.EnumMsg.SUCCESS)

    except Exception as e:
        if 'wb' in locals():
            wb.close()

        _, _, exc_tb = sys.exc_info()
        tb_info = traceback.extract_tb(exc_tb)[-1]

        error_msg = (
            f"Error: {e} | "
            f"File: {tb_info.filename} | "
            f"Line: {tb_info.lineno} | "
            f"Func: {tb_info.name}"
        )
        return uf.get_msg_struct(uf.EnumMsg.ERROR_SAVE_OR_PROC_TEMPLATE, error_msg)
# ============================================================================================
# ============================================================================================
def insert_sheet_key_in_tab(input_data_key):
    if input_data_key == None:
        return False

    db = uf.get_db_connection()

    # Переносим данные из словаря в выделенную память, используя табуляцию
    output = io.StringIO()
    try:
        for row in input_data_key:
            output.write(f'{row["sheet_id"]}\t{row["key_bs"]}\n')
        output.seek(0)
    except KeyError as e:
        return False

    try:
        with db.bind.begin() as conn:
            raw_conn = conn.connection.connection
            with raw_conn.cursor() as cursor:
                cursor.execute("""
                    CREATE TEMP TABLE temp_staging (
                        sheet_id int,
                        key_bs int
                    ) ON COMMIT DROP;
                """)

                cursor.copy_from(output, 'temp_staging', columns=('sheet_id', 'key_bs'))

                cursor.execute("""
                    INSERT INTO tab_sheet_input_data_d816_4
                        (sheet_id, key_bs)
                    SELECT
                        sheet_id,
                        key_bs
                    FROM temp_staging
                """)
        return True

    except errors.UniqueViolation as e:
        return False

    finally:
        # Освободить выделенную память
        output.close()

def insert_get_preu_mirror_in_tab(data):
    if data == None:
        return False

    db = uf.get_db_connection()

    # Переносим данные из словаря в выделенную память, используя табуляцию
    output = io.StringIO()
    try:
        for row in data:
            output.write('\t'.join(map(str, row.values())) + '\n')
        output.seek(0)
    except KeyError as e:
        return False

    try:
        first_row = data[0]
        # first_row_without_sum = [key for key in first_row.keys() if key != 'sum']
        first_row = [key for key in first_row.keys()]
        columns = ", ".join(first_row)
        columns_index = ", ".join(key for key in first_row if key != 'sum')
        with db.bind.begin() as conn:
            raw_conn = conn.connection.connection
            with raw_conn.cursor() as cursor:
                # Создаём временную табличку черновик со структурой целевой таблицы
                # через запрос с лимитом ноль получаем структуру
                cursor.execute("""
                    CREATE TEMP TABLE temp_staging ON COMMIT DROP AS 
                    SELECT * FROM tab_integ_get_preu_mirror_d816_4 LIMIT 0;
                """)

                cursor.copy_from(output, 'temp_staging', columns=tuple(first_row))

                cursor.execute(f"""
                    INSERT INTO tab_integ_get_preu_mirror_d816_4 ({columns})
                    SELECT {columns}
                    FROM temp_staging
                    ON CONFLICT ({columns_index}) 
                    DO UPDATE SET sum = EXCLUDED.sum;
                """)
        return True

    except errors.UniqueViolation as e:
        return False
    except Exception as e:
        return False

    finally:
        # Освободить выделенную память
        output.close()
